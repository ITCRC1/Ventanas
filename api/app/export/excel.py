"""Exportador del modelo Excel VIVO desde Postgres (Fase 8).

Porta la lógica de presentación del prototipo etl/export_excel.py — el docstring
del prototipo ya avisaba: "en producción cambia la consulta, no la lógica". Aquí
la fuente es la base de la app (Postgres) vía SQLAlchemy, no el SQLite del ETL.

Genera un .xlsx que:
  - se ve como el master (mismos colores/estados),
  - lleva FÓRMULAS reales (Spend, Remaining, subtotales, totales) — editás el
    LEDGER y todo recalcula,
  - agrupa de verdad en 3 niveles: semana -> mes -> trimestre.
"""

from __future__ import annotations

import contextlib
import datetime as dt
from collections import defaultdict
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter as gc
from openpyxl.worksheet.properties import Outline
from sqlalchemy import text
from sqlalchemy.engine.row import RowMapping
from sqlalchemy.orm import Session

STATES = [
    ("not_started", "Not Started", "CCCCCC"),
    ("in_process", "In process", "FFFF00"),
    ("approved", "Approved", "1155CC"),
    ("attention", "Attention", "B85B22"),
    ("completed", "Completed", "38761D"),
]
STATE_FILL = {c: PatternFill("solid", fgColor="FF" + h) for c, _, h in STATES}

HDR = PatternFill("solid", fgColor="FF434343")
SUBHDR = PatternFill("solid", fgColor="FFD9D9D9")
TOTAL = PatternFill("solid", fgColor="FFC8DCF0")
CATF = PatternFill("solid", fgColor="FFF8F8F8")
WHITE = Font(color="FFFFFFFF", bold=True)
BOLD = Font(bold=True)
MONEY = "#,##0.00;[Red](#,##0.00)"
DEFAULT_CAT_HEX = "94A3B8"  # gris para categorías sin color


def _soften(hex6: str) -> str:
    """Aclara un color 50% hacia blanco (igual que la app cuando la celda tiene
    cifra, para que se lea el número). hex6 = 'RRGGBB'."""
    h = hex6.lstrip("#")
    if len(h) != 6:
        return hex6
    r, g, b = (int(h[i : i + 2], 16) for i in (0, 2, 4))
    r, g, b = (round(c + (255 - c) * 0.5) for c in (r, g, b))
    return f"{r:02X}{g:02X}{b:02X}"


def _mondays(start: dt.date, end: dt.date) -> list[dt.date]:
    d = start - dt.timedelta(days=start.weekday())
    out: list[dt.date] = []
    while d <= end:
        out.append(d)
        d += dt.timedelta(days=7)
    return out


def _quarter(d: dt.date) -> str:
    return f"Q{(d.month - 1) // 3 + 1} {d.year}"


def _week_bounds(db: Session, start: dt.date | None, end: dt.date | None) -> tuple[dt.date, dt.date]:
    if start and end:
        return start, end
    row = db.execute(text("SELECT min(week_start), max(week_start) FROM schedule_cell")).first()
    lo = start or (row[0] if row and row[0] else dt.date(2026, 1, 5))
    hi = end or (row[1] if row and row[1] else dt.date(2026, 12, 28))
    return lo, hi


class ExcelExporter:
    """Construye el workbook leyendo de Postgres. Devuelve bytes (.xlsx)."""

    def __init__(self, db: Session, start: dt.date, end: dt.date) -> None:
        self.db = db
        self.weeks = _mondays(start, end)
        self.wb = Workbook()
        # Color por categoría (como el pincel del app: el color se comparte entre
        # Job Cost, Timeline y Timeline Detail porque salen del mismo schedule_cell).
        self.cat_hex: dict[str, str] = {
            (r["name"] or "").strip(): (r["color_hex"] or "").lstrip("#") or DEFAULT_CAT_HEX
            for r in db.execute(text("SELECT name, color_hex FROM category")).mappings()
        }

    def _cat_fill(self, category: str | None, has_amount: bool) -> PatternFill:
        """Relleno de una celda de semana PINTADA: color de la categoría (atenuado
        50% si tiene cifra, para leer el número) — idéntico en las 3 vistas."""
        hexc = self.cat_hex.get((category or "").strip(), DEFAULT_CAT_HEX)
        if has_amount:
            hexc = _soften(hexc)
        return PatternFill("solid", fgColor="FF" + hexc)


    def build(self) -> bytes:
        self._sheet_ledger()
        self._sheet_dashboard()
        self._sheet_jobcost()
        self._sheet_timeline()
        self._sheet_timeline_detail()
        self._sheet_short_payment()
        self._sheet_creditos()
        self._sheet_bancos()
        self._sheet_wires()
        self._sheet_facturas()
        self._sheet_legend()
        # Mismo orden que las pestañas del app.
        order = ["Timeline", "Timeline Detail", "Job Cost Report", "Dashboard",
                 "Short Payment", "LEDGER", "Credits", "Banks",
                 "US Wire Transfers", "Invoices", "Legend"]
        with contextlib.suppress(Exception):
            self.wb._sheets.sort(
                key=lambda sh: order.index(sh.title) if sh.title in order else 99
            )
        bio = BytesIO()
        self.wb.save(bio)
        return bio.getvalue()

    def _sheet_ledger(self) -> None:
        ws = self.wb.active
        ws.title = "LEDGER"
        cols = [
            "Cost Code", "Date", "Invoice #", "Payee", "Description",
            "Amount", "Amount Paid", "Amount Due", "Funding Source",
        ]
        ws.append([])
        ws.append([])
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            ws.cell(3, c).fill = HDR
            ws.cell(3, c).font = WHITE
        rows = self.db.execute(
            text(
                """
                SELECT w.wbs_code AS cost_code, l.entry_date AS date, l.invoice_no AS invoice,
                       p.name AS payee, l.description, l.amount, l.amount_paid, l.funding_source
                FROM ledger_entry l
                LEFT JOIN wbs_item w ON w.id = l.wbs_id
                LEFT JOIN payee p    ON p.id = l.payee_id
                ORDER BY l.id
                """
            )
        ).mappings()
        r = 4
        for e in rows:
            ws.cell(r, 1, (e["cost_code"] or "").strip())
            ws.cell(r, 2, e["date"])
            ws.cell(r, 3, e["invoice"])
            ws.cell(r, 4, e["payee"])
            ws.cell(r, 5, e["description"])
            ws.cell(r, 6, float(e["amount"] or 0)).number_format = MONEY
            ws.cell(r, 7, float(e["amount_paid"] or 0)).number_format = MONEY
            ws.cell(r, 8, f"=F{r}-G{r}").number_format = MONEY  # Amount Due VIVO
            ws.cell(r, 9, e["funding_source"])
            r += 1
        self.ledger_last = max(r - 1, 4)
        for c, w in zip(range(1, 10), [12, 12, 14, 26, 46, 14, 14, 14, 18], strict=False):
            ws.column_dimensions[gc(c)].width = w
        ws.freeze_panes = "B4"
        ws.auto_filter.ref = f"A3:I{self.ledger_last}"

    # ------------------------------------------------------------------ 2

    def _sheet_jobcost(self) -> None:
        """Job Cost Report FORMULADO — réplica de la vista del app (JobCostFull).

        Reemplaza el _sheet_jobcost anterior. Cambios clave vs. la base:
          - agrupa por CATEGORÍA (como el app: fila de categoría + subtotales);
          - por cada categoría, una fila TOTAL GASTOS (líneas kind<>'proceeds')
            y, SOLO si hay proceeds, una fila TOTAL VENTA (las líneas de venta
            NO se listan como filas: sólo suman en ese subtotal);
          - todo lo calculado es FÓRMULA VIVA (Revised=G+H, Spend=SUMIF al
            LEDGER, Remaining=I-J, Forecast=Spend+Timeline, Variance=I-L,
            subtotales con SUM de rango, TOTAL VENTA con SUMIF por código,
            PROJECT TOTAL = suma de las filas de subtotal — sin doble conteo);
          - mantiene semana->mes->trimestre con outline y el color del estado
            (schedule_cell.state_id) en cada celda de semana.
        Layout ANCLA intacto (cabecera fila 5): A=Status B=WBS C=Task Title
        D=Owner E=Category F=Phase G=Orig H=Change I=Revised J=Spend
        K=Remaining L=Forecast M=Variance, luego semana/mes/trim y TOTAL final.
        """
        ws = self.wb.create_sheet("Job Cost Report")
        ws.sheet_properties.outlinePr = Outline(
            summaryRight=True, summaryBelow=False, showOutlineSymbols=True
        )
        ws["A1"] = "DASHBOARD VENTANAS — Job Cost Report"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = "Live model. Edit the LEDGER and everything recalculates."
        ws["A2"].font = Font(italic=True, color="FF808080")

        GREEN = PatternFill("solid", fgColor="FFDFF3E3")  # fondo TOTAL VENTA
        WEEKFMT = "#,##0"

        fixed = [
            "Status", "WBS", "Task Title", "Owner", "Category", "Phase",
            "Budget Original", "Budget Change", "Budget Revised",
            "Spend", "Remaining", "Forecast", "Variance",
        ]
        hr = 5
        for i, h in enumerate(fixed, 1):
            c = ws.cell(hr, i, h)
            c.fill = HDR
            c.font = WHITE
            c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.row_dimensions[hr].height = 30

        # --- cabecera de tiempo: semana -> mes -> trimestre (con outline) ---
        col = len(fixed) + 1
        self.week_col: dict[dt.date, int] = {}
        self.month_col: dict[tuple[int, int], tuple[int, int, int]] = {}
        self.qtr_col: dict[str, tuple[int, list[int]]] = {}
        by_q: dict[str, dict[tuple[int, int], list[dt.date]]] = defaultdict(
            lambda: defaultdict(list)
        )
        for w in self.weeks:
            by_q[_quarter(w)][(w.year, w.month)].append(w)

        for qname, months in by_q.items():
            m_cols: list[int] = []
            for (yy, mm), wks in months.items():
                w_first = col
                for w in wks:
                    ws.cell(3, col, qname)
                    ws.cell(4, col, dt.date(yy, mm, 1).strftime("%b %Y").upper())
                    cc = ws.cell(hr, col, w.day)
                    cc.fill = SUBHDR
                    cc.font = Font(bold=True, size=9)
                    cc.alignment = Alignment(horizontal="center")
                    ws.column_dimensions[gc(col)].width = 5.5
                    self.week_col[w] = col
                    col += 1
                mc = ws.cell(hr, col, dt.date(yy, mm, 1).strftime("%b").upper())
                mc.fill = PatternFill("solid", fgColor="FFA2C4C9")
                mc.font = BOLD
                mc.alignment = Alignment(horizontal="center")
                ws.cell(3, col, qname)
                ws.cell(4, col, "MONTH")
                ws.column_dimensions[gc(col)].width = 12
                self.month_col[(yy, mm)] = (col, w_first, col - 1)
                m_cols.append(col)
                for k in range(w_first, col):
                    ws.column_dimensions[gc(k)].outlineLevel = 2
                col += 1
            qc = ws.cell(hr, col, qname)
            qc.fill = TOTAL
            qc.font = BOLD
            qc.alignment = Alignment(horizontal="center")
            ws.cell(3, col, qname)
            ws.cell(4, col, "QUARTER")
            ws.column_dimensions[gc(col)].width = 14
            self.qtr_col[qname] = (col, m_cols)
            for k in m_cols:
                ws.column_dimensions[gc(k)].outlineLevel = 1
            col += 1
        self.total_col = col
        tc = ws.cell(hr, col, "TOTAL")
        tc.fill = HDR
        tc.font = WHITE
        ws.column_dimensions[gc(col)].width = 15

        # Columnas de dinero de una fila (para subtotales/total): las 7 fijas
        # G..M + TODAS las de tiempo (semana/mes/trim) + la TOTAL final.
        money_cols = list(range(7, 14)) + list(range(14, self.total_col + 1))
        wcols = set(self.week_col.values())

        def numfmt(c: int) -> str:
            return WEEKFMT if c in wcols else MONEY

        # --- filas: v_wbs_financials, agrupadas por CATEGORÍA ---
        rows = self.db.execute(
            text(
                """
                SELECT f.wbs_code, f.title, f.owner, f.category, f.phase, f.kind,
                       ts.label AS status_label,
                       f.budget_original, f.budget_change,
                       w.forecast_total AS forecast_ovr
                FROM v_wbs_financials f
                LEFT JOIN wbs_item  w  ON w.wbs_code = f.wbs_code
                LEFT JOIN task_state ts ON ts.id = w.state_id
                WHERE f.kind <> 'section_header'
                ORDER BY COALESCE(f.category, 'zzz'), w.sort_order, f.wbs_code
                """
            )
        ).mappings().all()

        # Cronograma (schedule_cell): monto plan + estado por cost_code y semana.
        sched: defaultdict[str, dict[dt.date, RowMapping]] = defaultdict(dict)
        for s in self.db.execute(
            text(
                """
                SELECT w.wbs_code, s.week_start, s.planned_amount, ts.code AS state_code
                FROM schedule_cell s
                JOIN wbs_item w    ON w.id = s.wbs_id
                JOIN task_state ts ON ts.id = s.state_id
                """
            )
        ).mappings():
            sched[(s["wbs_code"] or "").strip()][s["week_start"]] = s

        # Agrupar líneas por categoría (orden que fija el ORDER BY de la query).
        cats: dict[str, list[RowMapping]] = defaultdict(list)
        for e in rows:
            cats[(e["category"] or "No category").strip()].append(e)

        def write_timeline_line(rr: int, code: str, category: str | None = None) -> None:
            """Celdas semana (valor base + COLOR DE LA CATEGORÍA, igual que el app —
            se sincroniza con Timeline/Detail) y mes/trim/total (fórmula)."""
            for w, wc in self.week_col.items():
                sc: RowMapping | None = sched[code].get(w)
                if not sc:
                    continue
                cell = ws.cell(rr, wc)
                amt = sc["planned_amount"]
                if amt:
                    cell.value = round(float(amt), 2)
                    cell.number_format = WEEKFMT
                cell.fill = self._cat_fill(category, bool(amt))
            for (_yy, _mm), (mc2, a, b) in self.month_col.items():
                ws.cell(rr, mc2, f"=SUM({gc(a)}{rr}:{gc(b)}{rr})").number_format = MONEY
            for _qn, (qc2, mcs) in self.qtr_col.items():
                ws.cell(rr, qc2, "=" + "+".join(f"{gc(m)}{rr}" for m in mcs)).number_format = MONEY
            ws.cell(
                rr, self.total_col,
                "=" + "+".join(f"{gc(q)}{rr}" for q, _ in self.qtr_col.values()),
            ).number_format = MONEY

        def write_timeline_venta(rr: int, codes: list[str]) -> None:
            """Igual, pero el valor semanal = suma de las líneas de venta (sin color)."""
            for w, wc in self.week_col.items():
                tot = 0.0
                for c in codes:
                    sc = sched[c].get(w)
                    if sc and sc["planned_amount"]:
                        tot += float(sc["planned_amount"])
                if tot:
                    ws.cell(rr, wc, round(tot, 2)).number_format = WEEKFMT
            for (_yy, _mm), (mc2, a, b) in self.month_col.items():
                ws.cell(rr, mc2, f"=SUM({gc(a)}{rr}:{gc(b)}{rr})").number_format = MONEY
            for _qn, (qc2, mcs) in self.qtr_col.items():
                ws.cell(rr, qc2, "=" + "+".join(f"{gc(m)}{rr}" for m in mcs)).number_format = MONEY
            ws.cell(
                rr, self.total_col,
                "=" + "+".join(f"{gc(q)}{rr}" for q, _ in self.qtr_col.values()),
            ).number_format = MONEY

        r = hr + 1
        subtotal_rows: list[int] = []  # filas TOTAL GASTOS/VENTA (para PROJECT TOTAL)

        for cat, lines in cats.items():
            costs = [e for e in lines if (e["kind"] or "cost") != "proceeds"]
            sales = [e for e in lines if (e["kind"] or "cost") == "proceeds"]

            # Encabezado de categoría (fila coloreada; título en col Task Title).
            for i in range(1, self.total_col + 1):
                ws.cell(r, i).fill = CATF
            ws.cell(r, 3, cat.upper()).font = BOLD
            r += 1

            # --- Líneas de GASTO (las de venta NO se listan) ---
            first_cost = r
            for e in costs:
                code = (e["wbs_code"] or "").strip()
                ws.cell(r, 1, e["status_label"])
                ws.cell(r, 2, code)
                ws.cell(r, 3, e["title"])
                ws.cell(r, 4, e["owner"])
                ws.cell(r, 5, cat)
                ws.cell(r, 6, e["phase"])
                ws.cell(r, 7, float(e["budget_original"] or 0)).number_format = MONEY  # base
                ws.cell(r, 8, float(e["budget_change"] or 0)).number_format = MONEY    # base
                # FÓRMULAS VIVAS
                ws.cell(r, 9, f"=G{r}+H{r}").number_format = MONEY                      # Revised
                ws.cell(
                    r, 10, f"=SUMIF(LEDGER!$A:$A,$B{r},LEDGER!$G:$G)"
                ).number_format = MONEY                                                # Spend
                ws.cell(r, 11, f"=I{r}-J{r}").number_format = MONEY                     # Remaining
                if e["forecast_ovr"] is not None:  # override manual del app = dato base
                    ws.cell(r, 12, float(e["forecast_ovr"])).number_format = MONEY
                else:  # Forecast = Spend + Timeline Total (vivo)
                    ws.cell(r, 12, f"=J{r}+{gc(self.total_col)}{r}").number_format = MONEY
                ws.cell(r, 13, f"=I{r}-L{r}").number_format = MONEY                     # Variance
                write_timeline_line(r, code, cat)
                r += 1
            last_cost = r - 1

            # --- TOTAL GASTOS (SUM del rango contiguo de líneas de gasto) ---
            for i in range(1, self.total_col + 1):
                ws.cell(r, i).fill = TOTAL
            ws.cell(r, 3, "TOTAL EXPENSES").font = BOLD
            for c in money_cols:
                if costs:
                    ws.cell(r, c, f"=SUM({gc(c)}{first_cost}:{gc(c)}{last_cost})")
                else:
                    ws.cell(r, c, 0)
                ws.cell(r, c).number_format = numfmt(c)
                ws.cell(r, c).font = BOLD
            subtotal_rows.append(r)
            r += 1

            # --- TOTAL VENTA (sólo si la categoría tiene proceeds) ---
            if sales:
                sale_codes = [(e["wbs_code"] or "").strip() for e in sales]
                for i in range(1, self.total_col + 1):
                    ws.cell(r, i).fill = GREEN
                ws.cell(r, 2, ", ".join(sale_codes))
                ws.cell(r, 3, "TOTAL SALES").font = BOLD
                # G/H = datos base (constantes); el resto, fórmula viva.
                ws.cell(
                    r, 7, sum(float(e["budget_original"] or 0) for e in sales)
                ).number_format = MONEY
                ws.cell(
                    r, 8, sum(float(e["budget_change"] or 0) for e in sales)
                ).number_format = MONEY
                ws.cell(r, 9, f"=G{r}+H{r}").number_format = MONEY
                spend_f = "+".join(
                    f'SUMIF(LEDGER!$A:$A,"{c}",LEDGER!$G:$G)' for c in sale_codes
                ) or "0"
                ws.cell(r, 10, f"={spend_f}").number_format = MONEY
                ws.cell(r, 11, f"=I{r}-J{r}").number_format = MONEY
                write_timeline_venta(r, sale_codes)
                ws.cell(r, 12, f"=J{r}+{gc(self.total_col)}{r}").number_format = MONEY
                ws.cell(r, 13, f"=I{r}-L{r}").number_format = MONEY
                for c in money_cols:
                    ws.cell(r, c).font = BOLD
                subtotal_rows.append(r)
                r += 1

        # --- PROJECT TOTAL (suma de las filas de subtotal; sin doble conteo) ---
        for i in range(1, self.total_col + 1):
            ws.cell(r, i).fill = TOTAL
        ws.cell(r, 3, "PROJECT TOTAL").font = BOLD
        if subtotal_rows:
            for c in money_cols:
                ws.cell(
                    r, c, "=" + "+".join(f"{gc(c)}{x}" for x in subtotal_rows)
                ).number_format = numfmt(c)
                ws.cell(r, c).font = BOLD

        for c, width in zip(
            range(1, 14), [11, 10, 42, 14, 20, 20, 15, 14, 15, 15, 15, 15, 14], strict=False
        ):
            ws.column_dimensions[gc(c)].width = width
        ws.freeze_panes = ws.cell(hr + 1, 7)
        ws.conditional_formatting.add(
            f"K{hr + 1}:K{r}",
            CellIsRule(
                operator="lessThan",
                formula=["0"],
                fill=PatternFill("solid", fgColor="FFF4CCCC"),
                font=Font(color="FF9C0006"),
            ),
        )

    def _sheet_timeline(self) -> None:
        """Hoja "Timeline": réplica FORMULADA de la vista del app (Timeline.tsx).

        Roll-up por categoría -> fase con columnas de la izquierda
        Budget Revised / YTD Expense / Remaining / Forecast / Funding Pending
        (fórmulas VIVAS al Job Cost Report) y, a la derecha, la matriz
        semana -> mes -> trimestre + FULL YEAR. Los montos de semana salen de
        schedule_cell (dato base = VALOR); mes/trimestre/año y subtotales son
        SUM/agregaciones vivas. El split GASTO/VENTA es por `kind`
        (kind='proceeds' = venta): la venta NO se lista como fila de fase, va
        sólo como subtotal TOTAL VENTA (idéntico al front).
        """
        JC = "'Job Cost Report'"
        MESES_UP = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG",
                    "SEP", "OCT", "NOV", "DEC"]
        VENTA_FILL = PatternFill("solid", fgColor="FFDFF3E3")
        GREY = Font(color="FF999999")

        def _clean(name: str) -> str:
            # Igual que cleanCatName del front: quita prefijo "N-" ("0-Property
            # Acquisition" -> "Property Acquisition"); NO toca "6.1 Overhead".
            i = name.find("-")
            if i > 0 and name[:i].strip().isdigit():
                return name[i + 1:].strip()
            return name

        ws = self.wb.create_sheet("Timeline")
        ws.sheet_properties.outlinePr = Outline(
            summaryRight=True, summaryBelow=False, showOutlineSymbols=True
        )
        ws["A1"] = "DASHBOARD VENTANAS — Timeline (roll-up by category → phase)"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = (
            "Live model. Budgets = formulas to the Job Cost Report; "
            "weeks = plan (schedule_cell). Edit the LEDGER/Job Cost and it recalculates."
        )
        ws["A2"].font = Font(italic=True, color="FF808080")

        # --- cabecera izquierda (7 columnas fijas, como LEFT_COLS del front) ---
        left = [
            "Project Category", "Phases", "Budget Revised", "YTD Expense",
            "Remaining", "Forecast", "Funding Pending",
        ]
        FIXED = len(left)  # 7
        hr = 5
        for i, h in enumerate(left, 1):
            c = ws.cell(hr, i, h)
            c.fill = HDR
            c.font = WHITE
            c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.row_dimensions[hr].height = 30

        # --- matriz semana -> mes -> trimestre (mismo layout que Job Cost) ---
        col = FIXED + 1
        week_col: dict[dt.date, int] = {}
        month_col: dict[tuple[int, int], tuple[int, int, int]] = {}
        qtr_col: dict[str, tuple[int, list[int]]] = {}
        by_q: dict[str, dict[tuple[int, int], list[dt.date]]] = defaultdict(
            lambda: defaultdict(list)
        )
        for w in self.weeks:
            by_q[_quarter(w)][(w.year, w.month)].append(w)

        for qname, months in by_q.items():
            m_cols: list[int] = []
            for (yy, mm), wks in months.items():
                w_first = col
                for w in wks:
                    ws.cell(3, col, qname)
                    ws.cell(4, col, f"{MESES_UP[mm - 1]} {yy}")
                    cc = ws.cell(hr, col, w.day)
                    cc.fill = SUBHDR
                    cc.font = Font(bold=True, size=9)
                    cc.alignment = Alignment(horizontal="center")
                    ws.column_dimensions[gc(col)].width = 5.5
                    ws.column_dimensions[gc(col)].outlineLevel = 2
                    week_col[w] = col
                    col += 1
                mc = ws.cell(hr, col, MESES_UP[mm - 1])
                mc.fill = PatternFill("solid", fgColor="FFA2C4C9")
                mc.font = BOLD
                mc.alignment = Alignment(horizontal="center")
                ws.cell(3, col, qname)
                ws.cell(4, col, "MONTH")
                ws.column_dimensions[gc(col)].width = 12
                ws.column_dimensions[gc(col)].outlineLevel = 1
                month_col[(yy, mm)] = (col, w_first, col - 1)
                m_cols.append(col)
                col += 1
            qc = ws.cell(hr, col, qname)
            qc.fill = TOTAL
            qc.font = BOLD
            qc.alignment = Alignment(horizontal="center")
            ws.cell(3, col, qname)
            ws.cell(4, col, "QUARTER")
            ws.column_dimensions[gc(col)].width = 14
            qtr_col[qname] = (col, m_cols)
            col += 1
        total_col = col
        tc = ws.cell(hr, col, "FULL YEAR")
        tc.fill = PatternFill("solid", fgColor="FF1A7F4B")
        tc.font = WHITE
        tc.alignment = Alignment(horizontal="center")
        ws.column_dimensions[gc(col)].width = 15

        week_cols_all = list(week_col.values())
        month_cols_all = [c for c, _, _ in month_col.values()]
        qtr_cols_all = [c for c, _ in qtr_col.values()]
        week_set = set(week_cols_all)
        agg_grid = week_cols_all + month_cols_all + qtr_cols_all + [total_col]

        # --- datos: líneas (v_wbs_financials, orden wbs_code igual que la API) ---
        lines = self.db.execute(
            text(
                """
                SELECT f.id, f.wbs_code, f.category, f.phase, f.kind
                FROM v_wbs_financials f
                WHERE f.kind <> 'section_header'
                ORDER BY f.wbs_code
                """
            )
        ).mappings().all()
        sched: dict[tuple[int, dt.date], float] = {}
        for s in self.db.execute(
            text("SELECT wbs_id, week_start, planned_amount FROM schedule_cell")
        ).mappings():
            # Guardamos TODAS las celdas (presencia = pintada), monto 0 si es sólo color.
            sched[(s["wbs_id"], s["week_start"])] = float(s["planned_amount"] or 0)
        # orden de categorías = category.sort_order (como tool.catOrder del front),
        # incluyendo las vacías (aparecen con su TOTAL en 0, igual que el app).
        cat_order = self.db.execute(
            text("SELECT name FROM category WHERE is_active ORDER BY sort_order, name")
        ).scalars().all()

        # agrupar categoría -> fase (costos) + ventas (proceeds), preservando el
        # orden de wbs_code para las fases (idéntico al Map del front).
        groups: dict[str, dict[str, Any]] = {}
        for e in lines:
            cat = e["category"] or "No category"
            g = groups.setdefault(cat, {"phases": {}, "sales": []})
            code = (e["wbs_code"] or "").strip()
            if e["kind"] == "proceeds":
                g["sales"].append((e["id"], code))
            else:
                ph = e["phase"] or "No phase"
                g["phases"].setdefault(ph, []).append((e["id"], code))
        ordered = list(cat_order)
        for n in groups:
            if n not in ordered:
                ordered.append(n)

        # ---- helpers de fórmulas ----
        def _sumif(codes: list[tuple[int, str]], src: str) -> object:
            # Presupuesto/gasto/forecast de un conjunto de WBS: SUMIF al Job Cost.
            if not codes:
                return 0
            return "=" + "+".join(
                f'SUMIF({JC}!$B:$B,"{c}",{JC}!${src}:${src})' for _i, c in codes
            )

        def _leaf(
            row: int,
            codes: list[tuple[int, str]],
            bold: bool = False,
            category: str | None = None,
            paint: bool = True,
        ) -> None:
            # Fila "hoja" (fase de costos o subtotal de venta): budgets vivos al
            # Job Cost + semanas por VALOR (schedule_cell), coloreadas por CATEGORÍA
            # (igual que Job Cost/Detail) + mes/qtr/año por SUM.
            ws.cell(row, 3, _sumif(codes, "I")).number_format = MONEY   # Budget Revised
            ws.cell(row, 4, _sumif(codes, "J")).number_format = MONEY   # YTD Expense (spend)
            ws.cell(row, 5, f"={gc(3)}{row}-{gc(4)}{row}").number_format = MONEY  # Remaining
            ws.cell(row, 6, _sumif(codes, "L")).number_format = MONEY   # Forecast
            ws.cell(row, 7, f"={gc(6)}{row}-{gc(4)}{row}").number_format = MONEY  # Funding Pending
            ids = [i for i, _ in codes]
            for w, wc in week_col.items():
                if not any((i, w) in sched for i in ids):
                    continue
                v = sum(sched.get((i, w), 0.0) for i in ids)
                cell = ws.cell(row, wc)
                if v:
                    cell.value = round(v, 2)
                    cell.number_format = "#,##0"
                if paint:
                    cell.fill = self._cat_fill(category, bool(v))
            for (_yy, _mm), (mc2, a, b) in month_col.items():
                ws.cell(row, mc2, f"=SUM({gc(a)}{row}:{gc(b)}{row})").number_format = MONEY
            for _qn, (qc2, mcs) in qtr_col.items():
                ws.cell(row, qc2, "=" + "+".join(f"{gc(m)}{row}" for m in mcs)).number_format = MONEY
            ws.cell(
                row, total_col,
                "=" + "+".join(f"{gc(q)}{row}" for q, _ in qtr_col.values()),
            ).number_format = MONEY
            if bold:
                for cc2 in range(3, 8):
                    ws.cell(row, cc2).font = BOLD

        def _agg(row: int, srcs: list[int], label: str, fill: PatternFill) -> None:
            # Fila agregada (TOTAL GASTOS / TOTAL <cat> / PROJECT TOTAL): cada
            # columna = suma VIVA de las filas fuente (fases o subtotales).
            for i in range(1, total_col + 1):
                ws.cell(row, i).fill = fill
            ws.cell(row, 1, label).font = BOLD

            def _s(c: int) -> object:
                return ("=" + "+".join(f"{gc(c)}{rr}" for rr in srcs)) if srcs else 0

            ws.cell(row, 3, _s(3)).number_format = MONEY
            ws.cell(row, 4, _s(4)).number_format = MONEY
            ws.cell(row, 5, f"={gc(3)}{row}-{gc(4)}{row}").number_format = MONEY
            ws.cell(row, 6, _s(6)).number_format = MONEY
            ws.cell(row, 7, f"={gc(6)}{row}-{gc(4)}{row}").number_format = MONEY
            for c in agg_grid:
                cell = ws.cell(row, c, _s(c))
                cell.number_format = "#,##0" if c in week_set else MONEY
            for c in range(3, 8):
                ws.cell(row, c).font = BOLD
            for c in agg_grid:
                ws.cell(row, c).font = BOLD

        # ---- filas ----
        r = hr + 1
        subtotal_rows: list[int] = []
        for cat in ordered:
            g = groups[cat]
            # encabezado de categoría (banda)
            for i in range(1, total_col + 1):
                ws.cell(r, i).fill = CATF
            ws.cell(r, 1, _clean(cat).upper()).font = BOLD
            r += 1
            # filas de fase (sólo costos; la venta va aparte)
            phase_rows: list[int] = []
            sales: list[tuple[int, str]] = []
            if g:
                for ph, codes in g["phases"].items():
                    ws.cell(r, 1, _clean(cat)).font = GREY
                    ws.cell(r, 2, ph)
                    _leaf(r, codes, category=cat)
                    phase_rows.append(r)
                    r += 1
                sales = g["sales"]
            # subtotales por categoría (misma regla de kind que el front)
            if sales:
                _agg(r, phase_rows, "TOTAL EXPENSES", TOTAL)
                subtotal_rows.append(r)
                r += 1
                for i in range(1, total_col + 1):
                    ws.cell(r, i).fill = VENTA_FILL
                ws.cell(r, 1, "TOTAL SALES · " + ", ".join(c for _i, c in sales)).font = BOLD
                _leaf(r, sales, bold=True, paint=False)  # mantiene el verde de la fila
                subtotal_rows.append(r)
                r += 1
            else:
                _agg(r, phase_rows, f"TOTAL {_clean(cat).upper()}", TOTAL)
                subtotal_rows.append(r)
                r += 1

        # PROJECT TOTAL = suma de todos los subtotales (= todas las líneas,
        # costos + ventas, igual que `grand` del front).
        _agg(r, subtotal_rows, "PROJECT TOTAL", TOTAL)

        # ancho de columnas izquierdas + congelar identidad/presupuesto.
        for c, wd in zip(range(1, 8), [22, 26, 15, 14, 14, 14, 15], strict=False):
            ws.column_dimensions[gc(c)].width = wd
        ws.freeze_panes = ws.cell(hr + 1, FIXED + 1)

    def _sheet_timeline_detail(self) -> None:
        """Timeline Detail — réplica FORMULADA de la vista del app (TimelineDetail.tsx).

        Mismas columnas de identidad + presupuesto (Project Category / Phases /
        Phases Subphase / WBS Number / Task Title / Owner / Budget Revised / YTD
        Expense / Remaining / Forecast / Funding Pending), las semanas agrupadas
        semana→mes→trimestre y un TOTAL AÑO, con las categorías agrupadas
        (cabecera + subtotal), en el orden del Excel (category.sort_order).

        Donde el app CALCULA, hay FÓRMULA VIVA: Budget Revised / YTD / Remaining /
        Forecast se traen del 'Job Cost Report' por SUMIF sobre el WBS; Funding =
        Forecast − YTD; los montos por semana son el dato base (planned_amount) y
        meses/trimestres/totales/subtotales son SUM de esas celdas.

        Split de venta: la línea kind='proceeds' NO se lista; aparece SOLO como
        fila de subtotal 'TOTAL VENTA · <wbs>'. Las líneas de gasto sí se listan y
        suman en 'TOTAL GASTOS' (o 'TOTAL <cat>' si la categoría no tiene venta).
        El PROJECT TOTAL suma TODAS las líneas (gastos + venta), como el app.
        """
        ws = self.wb.create_sheet("Timeline Detail")
        ws.sheet_properties.outlinePr = Outline(
            summaryRight=True, summaryBelow=False, showOutlineSymbols=True
        )
        # Fills propios de esta vista (colores del app; los de módulo se reutilizan)
        CAT_FILL = PatternFill("solid", fgColor="FFEAF0FA")   # cabecera categoría
        GASTO_FILL = PatternFill("solid", fgColor="FFDCE6F4")  # TOTAL / TOTAL GASTOS
        VENTA_FILL = PatternFill("solid", fgColor="FFDFF3E3")  # TOTAL VENTA (verde)
        RED_FILL = PatternFill("solid", fgColor="FFF4CCCC")
        RED_FONT = Font(color="FF9C0006")
        WK = "#,##0"

        # ---- columnas fijas: identidad + presupuesto (idénticas al app) ----
        fixed = [
            "Project Category", "Phases", "Phases Subphase", "WBS Number",
            "Task Title", "Owner", "Budget Revised", "YTD Expense",
            "Remaining", "Forecast", "Funding Pending",
        ]
        FIX = len(fixed)  # 11
        hr = 5            # cabecera en fila 5 (3 filas de tiempo, como Job Cost)
        for i, h in enumerate(fixed, 1):
            c = ws.cell(hr, i, h)
            c.fill = HDR
            c.font = WHITE
            c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.row_dimensions[hr].height = 30

        # ---- columnas de tiempo: semana -> mes -> trimestre (como Job Cost) ----
        col = FIX + 1
        week_col: dict[dt.date, int] = {}
        month_col: dict[tuple[int, int], tuple[int, int, int]] = {}
        qtr_col: dict[str, tuple[int, list[int]]] = {}
        by_q: dict[str, dict[tuple[int, int], list[dt.date]]] = defaultdict(
            lambda: defaultdict(list)
        )
        for w in self.weeks:
            by_q[_quarter(w)][(w.year, w.month)].append(w)
        for qname, months in by_q.items():
            m_cols: list[int] = []
            for (yy, mm), wks in months.items():
                w_first = col
                for w in wks:
                    ws.cell(3, col, qname)
                    ws.cell(4, col, dt.date(yy, mm, 1).strftime("%b %Y").upper())
                    cc = ws.cell(hr, col, w.day)
                    cc.fill = SUBHDR
                    cc.font = Font(bold=True, size=9)
                    cc.alignment = Alignment(horizontal="center")
                    ws.column_dimensions[gc(col)].width = 5.5
                    week_col[w] = col
                    col += 1
                mc = ws.cell(hr, col, dt.date(yy, mm, 1).strftime("%b").upper())
                mc.fill = PatternFill("solid", fgColor="FFA2C4C9")
                mc.font = BOLD
                mc.alignment = Alignment(horizontal="center")
                ws.cell(3, col, qname)
                ws.cell(4, col, "MONTH")
                ws.column_dimensions[gc(col)].width = 11
                month_col[(yy, mm)] = (col, w_first, col - 1)
                m_cols.append(col)
                for k in range(w_first, col):
                    ws.column_dimensions[gc(k)].outlineLevel = 2
                col += 1
            qc = ws.cell(hr, col, qname)
            qc.fill = TOTAL
            qc.font = BOLD
            qc.alignment = Alignment(horizontal="center")
            ws.cell(3, col, qname)
            ws.cell(4, col, "QUARTER")
            ws.column_dimensions[gc(col)].width = 13
            qtr_col[qname] = (col, m_cols)
            for k in m_cols:
                ws.column_dimensions[gc(k)].outlineLevel = 1
            col += 1
        total_col = col
        tc = ws.cell(hr, col, "FULL YEAR")
        tc.fill = HDR
        tc.font = WHITE
        tc.alignment = Alignment(horizontal="center")
        ws.column_dimensions[gc(col)].width = 15

        # ---- rollup de tiempo de una fila (mes=SUM semanas, trim=SUM meses, total=SUM trim) ----
        def _rollup(row: int) -> None:
            for (_yy, _mm), (mc2, a, b) in month_col.items():
                ws.cell(row, mc2, f"=SUM({gc(a)}{row}:{gc(b)}{row})").number_format = MONEY
            for _qn, (qc2, mcs) in qtr_col.items():
                ws.cell(
                    row, qc2, "=" + "+".join(f"{gc(m)}{row}" for m in mcs)
                ).number_format = MONEY
            ws.cell(
                row, total_col,
                "=" + "+".join(f"{gc(q)}{row}" for q, _ in qtr_col.values()),
            ).number_format = MONEY

        # ---- datos: líneas de v_wbs_financials + montos por semana (dato base) ----
        rows = self.db.execute(
            text(
                """
                SELECT f.wbs_code, f.title, f.owner, f.category, f.phase, f.kind
                FROM v_wbs_financials f
                LEFT JOIN wbs_item w ON w.wbs_code = f.wbs_code
                LEFT JOIN category c ON c.id = w.category_id
                WHERE f.kind <> 'section_header'
                ORDER BY COALESCE(c.sort_order, 999),
                         COALESCE(f.category, 'zzz'), f.wbs_code
                """
            )
        ).mappings().all()
        sched: defaultdict[str, dict[dt.date, float]] = defaultdict(dict)
        for s in self.db.execute(
            text(
                """
                SELECT w.wbs_code, s.week_start, s.planned_amount
                FROM schedule_cell s
                JOIN wbs_item w ON w.id = s.wbs_id
                """
            )
        ).mappings():
            sched[(s["wbs_code"] or "").strip()][s["week_start"]] = float(
                s["planned_amount"] or 0
            )

        # Agrupar por categoría en el orden del Excel; sembrar TODAS las categorías
        # activas (incluidas las vacías, ej. Construction) como hace el app.
        groups: dict[str, list[RowMapping]] = {}
        for name in self.db.execute(
            text("SELECT name FROM category WHERE is_active ORDER BY sort_order, code")
        ).scalars():
            groups[(name or "").strip()] = []
        for e in rows:
            cat = (e["category"] or "No category").strip()
            groups.setdefault(cat, []).append(e)

        # ---- helpers de escritura ----
        def _cost_row(row: int, e: RowMapping) -> None:
            code = (e["wbs_code"] or "").strip()
            ws.cell(row, 1, (e["category"] or "").strip())
            ws.cell(row, 2, e["phase"])
            ws.cell(row, 3, "—")            # Phases Subphase (el app muestra "—")
            ws.cell(row, 4, code)
            ws.cell(row, 5, e["title"])
            ws.cell(row, 6, e["owner"])
            # Budget Revised / YTD / Remaining / Forecast desde Job Cost por WBS
            for i, src in zip(range(7, 11), "IJKL", strict=False):
                ws.cell(
                    row, i,
                    f"=SUMIF('Job Cost Report'!$B:$B,$D{row},"
                    f"'Job Cost Report'!${src}:${src})",
                ).number_format = MONEY
            ws.cell(row, 11, f"=J{row}-H{row}").number_format = MONEY  # Funding = FC-YTD
            row_sched = sched.get(code, {})
            for w, wc in week_col.items():
                if w not in row_sched:
                    continue
                v = row_sched[w]
                cell = ws.cell(row, wc)
                if v:
                    cell.value = round(v, 2)
                    cell.number_format = WK
                cell.fill = self._cat_fill(e["category"], bool(v))  # color de categoría, igual que el app
            _rollup(row)

        def _venta_row(row: int, label: str, codes: list[str]) -> None:
            for i, src in zip(range(7, 11), "IJKL", strict=False):
                ws.cell(
                    row, i,
                    "=" + "+".join(
                        f"SUMIF('Job Cost Report'!$B:$B,\"{cd}\","
                        f"'Job Cost Report'!${src}:${src})"
                        for cd in codes
                    ),
                ).number_format = MONEY
            ws.cell(row, 11, f"=J{row}-H{row}").number_format = MONEY
            for w, wc in week_col.items():
                v = sum(sched.get(cd, {}).get(w, 0.0) for cd in codes)
                if v:
                    ws.cell(row, wc, round(v, 2)).number_format = WK
            _rollup(row)
            lab = ws.cell(row, 1, label)
            lab.font = BOLD
            for i in range(1, total_col + 1):
                ws.cell(row, i).fill = VENTA_FILL

        def _sum_row(row: int, label: str, idxs: list[int], fill: PatternFill) -> None:
            for i in range(7, 11):
                ws.cell(
                    row, i,
                    ("=" + "+".join(f"{gc(i)}{x}" for x in idxs)) if idxs else 0,
                ).number_format = MONEY
            ws.cell(row, 11, f"=J{row}-H{row}").number_format = MONEY
            for wc in week_col.values():
                if idxs:
                    ws.cell(
                        row, wc, "=" + "+".join(f"{gc(wc)}{x}" for x in idxs)
                    ).number_format = WK
            _rollup(row)
            lab = ws.cell(row, 1, label)
            lab.font = BOLD
            for i in range(1, total_col + 1):
                ws.cell(row, i).fill = fill

        # ---- cuerpo: categoría (cabecera + gastos + subtotal[s]) ----
        r = hr + 1
        subtotal_rows: list[int] = []
        for cat, lines in groups.items():
            for i in range(1, total_col + 1):
                ws.cell(r, i).fill = CAT_FILL
            ws.cell(r, 1, cat.upper()).font = BOLD
            r += 1

            costs = [e for e in lines if e["kind"] != "proceeds"]
            sales = [e for e in lines if e["kind"] == "proceeds"]

            cost_idx: list[int] = []
            for e in costs:
                _cost_row(r, e)
                cost_idx.append(r)
                r += 1

            if sales:
                _sum_row(r, "TOTAL EXPENSES", cost_idx, GASTO_FILL)
                subtotal_rows.append(r)
                r += 1
                sale_codes = [(e["wbs_code"] or "").strip() for e in sales]
                _venta_row(r, "TOTAL SALES · " + ", ".join(sale_codes), sale_codes)
                subtotal_rows.append(r)
                r += 1
            else:
                _sum_row(r, f"TOTAL {cat.upper()}", cost_idx, GASTO_FILL)
                subtotal_rows.append(r)
                r += 1

        # ---- PROJECT TOTAL: suma TODAS las líneas (= suma de todos los subtotales) ----
        _sum_row(r, "PROJECT TOTAL", subtotal_rows, TOTAL)
        last = r
        r += 1

        # ---- Remaining / Funding negativos en rojo (como el app) ----
        for col_l in ("I", "K"):
            ws.conditional_formatting.add(
                f"{col_l}{hr + 1}:{col_l}{last}",
                CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL, font=RED_FONT),
            )

        for c, wdt in zip(
            range(1, FIX + 1), [16, 12, 12, 12, 42, 14, 14, 13, 13, 13, 14], strict=False
        ):
            ws.column_dimensions[gc(c)].width = wdt
        ws.freeze_panes = ws.cell(hr + 1, FIX + 1)

    # ------------------------------------------------------------------ 5
    def _sheet_dashboard(self) -> None:
        """Réplica FORMULADA de la vista Dashboard del app.

        KPIs (Presupuesto revisado / Gasto a la fecha / Remanente / % ejecutado)
        y tabla por categoría con TOTAL GASTOS / TOTAL VENTA (kind='proceeds') y
        GRAN TOTAL. TODO por fórmula viva referenciando 'Job Cost Report':
          * Presupuesto = SUMIF(...$B:$B, wbs, ...$I:$I)   (Budget Revised)
          * Gasto       = SUMIF(...$B:$B, wbs, ...$J:$J)   (Spend = SUMIF al LEDGER)
          * Remanente   = SUMIF(...$B:$B, wbs, ...$K:$K)   (Remaining)
        Los subtotales/GRAN TOTAL suman celdas vivas; el % = Gasto/Presupuesto.
        La línea de venta NO se lista: aparece sólo como TOTAL VENTA.
        """
        ws = self.wb.create_sheet("Dashboard")
        PCT = "0.0%;[Red](0.0%)"
        VENTA_FILL = PatternFill("solid", fgColor="FFDFF3E3")   # verde venta
        VENTA_FONT = Font(bold=True, color="FF38761D")
        KPI_LBL = Font(size=9, color="FF808080")
        KPI_VAL = Font(bold=True, size=16)
        JC = "'Job Cost Report'"

        def sumif(code: str, col: str) -> str:
            c = (code or "").replace('"', '""')
            return f'SUMIF({JC}!$B:$B,"{c}",{JC}!${col}:${col})'

        ws["A1"] = "DASHBOARD VENTANAS — Summary"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = "Live model. Edit the LEDGER / Job Cost and everything recalculates."
        ws["A2"].font = Font(italic=True, color="FF808080")

        # --- datos (v_wbs_financials, mismo origen que Job Cost) ------------
        rows = self.db.execute(
            text(
                """
                SELECT f.wbs_code, f.title, f.state, f.kind, f.category,
                       f.budget_original, f.budget_change
                FROM v_wbs_financials f
                LEFT JOIN wbs_item w ON w.wbs_code = f.wbs_code
                WHERE f.kind <> 'section_header'
                ORDER BY COALESCE(f.category, 'zzz'), w.sort_order, f.wbs_code
                """
            )
        ).mappings().all()

        # agrupar por categoría preservando el orden de llegada de las líneas
        groups: dict[str, list[RowMapping]] = defaultdict(list)
        for e in rows:
            groups[(e["category"] or "No category").strip()].append(e)
        # el app ordena las categorías por presupuesto revisado descendente
        def cat_budget(lines: list[RowMapping]) -> float:
            return sum(
                float(ln["budget_original"] or 0) + float(ln["budget_change"] or 0)
                for ln in lines
            )
        ordered = sorted(groups.items(), key=lambda kv: cat_budget(kv[1]), reverse=True)

        # --- cabecera de la tabla ------------------------------------------
        hr = 7
        heads = ["Category / line", "Status", "Budget", "Spend", "Remaining"]
        for i, h in enumerate(heads, 1):
            c = ws.cell(hr, i, h)
            c.fill = HDR
            c.font = WHITE
            c.alignment = Alignment(
                horizontal="left" if i == 1 else ("center" if i == 2 else "right")
            )

        total_rows: list[int] = []   # filas de subtotal → suman al GRAN TOTAL
        r = hr + 1
        for cat, lines in ordered:
            # fila de categoría
            for i in range(1, 6):
                ws.cell(r, i).fill = CATF
            cc = ws.cell(r, 1, cat.upper())
            cc.font = BOLD
            r += 1

            costs = [ln for ln in lines if ln["kind"] != "proceeds"]
            sales = [ln for ln in lines if ln["kind"] == "proceeds"]

            cost_line_rows: list[int] = []
            for ln in costs:
                code = (ln["wbs_code"] or "").strip()
                ws.cell(r, 1, ln["title"]).alignment = Alignment(indent=1)
                st = ws.cell(r, 2, ln["state"])
                st.fill = STATE_FILL.get(ln["state"], STATE_FILL["not_started"])
                st.alignment = Alignment(horizontal="center")
                ws.cell(r, 3, f"={sumif(code, 'I')}").number_format = MONEY
                ws.cell(r, 4, f"={sumif(code, 'J')}").number_format = MONEY
                ws.cell(r, 5, f"={sumif(code, 'K')}").number_format = MONEY
                cost_line_rows.append(r)
                r += 1

            def sum_local(col: int, rr: list[int]) -> str:
                return "=" + "+".join(f"{gc(col)}{x}" for x in rr) if rr else "=0"

            if not sales:
                # una sola fila TOTAL (= suma de las líneas listadas)
                for i in range(1, 6):
                    ws.cell(r, i).fill = TOTAL
                ws.cell(r, 2, "TOTAL").font = BOLD
                ws.cell(r, 2).alignment = Alignment(horizontal="right")
                for col in (3, 4, 5):
                    ws.cell(r, col, sum_local(col, cost_line_rows)).number_format = MONEY
                    ws.cell(r, col).font = BOLD
                total_rows.append(r)
                r += 1
            else:
                # TOTAL GASTOS (suma de las líneas de gasto listadas)
                for i in range(1, 6):
                    ws.cell(r, i).fill = TOTAL
                ws.cell(r, 2, "TOTAL EXPENSES").font = BOLD
                ws.cell(r, 2).alignment = Alignment(horizontal="right")
                for col in (3, 4, 5):
                    ws.cell(r, col, sum_local(col, cost_line_rows)).number_format = MONEY
                    ws.cell(r, col).font = BOLD
                total_rows.append(r)
                r += 1
                # TOTAL VENTA (kind='proceeds') — no se lista, sólo el subtotal
                codes = ", ".join((ln["wbs_code"] or "").strip() for ln in sales)
                for i in range(1, 6):
                    ws.cell(r, i).fill = VENTA_FILL
                ws.cell(r, 2, f"TOTAL SALES · {codes}").font = VENTA_FONT
                ws.cell(r, 2).alignment = Alignment(horizontal="right")
                for col, src in ((3, "I"), (4, "J"), (5, "K")):
                    f = "=" + "+".join(
                        sumif((ln["wbs_code"] or "").strip(), src) for ln in sales
                    )
                    ws.cell(r, col, f).number_format = MONEY
                    ws.cell(r, col).font = VENTA_FONT
                total_rows.append(r)
                r += 1

        # --- GRAN TOTAL (suma de todas las filas de subtotal) --------------
        for i in range(1, 6):
            ws.cell(r, i).fill = TOTAL
        gt = ws.cell(r, 1, "GRAND TOTAL")
        gt.font = Font(bold=True, size=11)
        for col in (3, 4, 5):
            f = "=" + "+".join(f"{gc(col)}{x}" for x in total_rows) if total_rows else "=0"
            ws.cell(r, col, f).number_format = MONEY
            ws.cell(r, col).font = Font(bold=True, size=11)
        gt_row = r

        # --- KPIs (referencian el GRAN TOTAL, fórmulas vivas) --------------
        kpis = [
            (1, "Revised Budget", f"=C{gt_row}", MONEY),
            (2, "Spend to Date", f"=D{gt_row}", MONEY),
            (3, "Remaining", f"=E{gt_row}", MONEY),
            (4, "% Executed", f"=IFERROR(D{gt_row}/C{gt_row},0)", PCT),
        ]
        for col, lbl, formula, fmt in kpis:
            lc = ws.cell(4, col, lbl)
            lc.font = KPI_LBL
            vc = ws.cell(5, col, formula)
            vc.number_format = fmt
            vc.font = KPI_VAL

        for c, w in zip(range(1, 6), [40, 16, 16, 16, 16], strict=False):
            ws.column_dimensions[gc(c)].width = w
        ws.freeze_panes = ws.cell(hr + 1, 1)

    def _sheet_short_payment(self) -> None:
        """Short Payment List — RÉPLICA de ShortPaymentsView.tsx, pero formulada.

        Una sección por tanda (cada disbursement): barra de título
        'To be sent <fecha> · Desembolso #<no>.<sub> · <estado>', fila de
        encabezados de columna, las líneas del desembolso y una fila TOTAL con
        SUM viva sobre la columna Monto. El total de la barra de título es una
        referencia viva a esa misma suma.

        Columnas EXACTAS del app (LineRow):
          #, WBS, Descripción, Factura #, Nota, Categoría / Tipo, Nombre,
          Transfer, [Banco, Beneficiario, IBAN / Cuenta], Monto
        Las 3 columnas bancarias solo aparecen si algún renglón trae banco/IBAN
        (misma regla `showBank` del front). El delete-column de la vista no se
        exporta.
        """
        ws = self.wb.create_sheet("Short Payment")

        # STATUS_LABEL del front (ShortPaymentsView.tsx)
        STATUS_LABEL = {
            "draft": "Draft",
            "submitted": "Submitted to corporate",
            "approved": "Approved",
            "funded": "Wire received",
            "transferred": "Transferred to LAFISE",
            "settled": "Settled",
            "cancelled": "Cancelled",
        }
        MESES = [
            "", "January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December",
        ]

        # Misma consulta que repo.short_payments(): WBS/Categoría/Tipo se
        # resuelven por línea y se heredan del WBS si la línea no los trae.
        rows = self.db.execute(
            text(
                """
                SELECT d.id, d.disb_no, d.disb_sub, d.period_month, d.send_date, d.status,
                       d.total_amount,
                       dl.id AS line_id, dl.line_no, dl.description, dl.invoice_no,
                       dl.amount, dl.reason, dl.transfer,
                       w.wbs_code, w.title AS wbs_title,
                       COALESCE(cl.name, cw.name) AS category,
                       COALESCE(pl.name, pw.name) AS type,
                       p.name AS payee_name, p.bank_name, p.legal_id, p.iban
                FROM disbursement d
                LEFT JOIN disbursement_line dl ON dl.disbursement_id = d.id
                LEFT JOIN payee p    ON p.id  = dl.payee_id
                LEFT JOIN wbs_item w ON w.id  = dl.wbs_id
                LEFT JOIN category cl ON cl.id = dl.category_id
                LEFT JOIN phase    pl ON pl.id = dl.phase_id
                LEFT JOIN category cw ON cw.id = w.category_id
                LEFT JOIN phase    pw ON pw.id = w.phase_id
                ORDER BY d.disb_no, d.disb_sub, dl.line_no
                """
            )
        ).mappings().all()

        if not rows:
            ws["A1"] = (
                "No payment batches yet. They are created from the "
                "Disbursements tab (or when importing from Excel)."
            )
            ws["A1"].font = Font(italic=True, color="FF808080")
            ws.column_dimensions["A"].width = 80
            return

        # showBank: igual que el front (algún renglón con banco o IBAN)
        show_bank = any((r["bank_name"] or r["iban"]) for r in rows)

        base = ["#", "WBS", "Description", "Invoice #", "Note",
                "Category / Type", "Name", "Transfer"]
        if show_bank:
            base += ["Bank", "Beneficiary", "IBAN / Account"]
        base += ["Amount"]
        monto_col = len(base)           # última columna
        col_letter = gc(monto_col)

        widths = {
            "#": 5, "WBS": 12, "Description": 42, "Invoice #": 12, "Note": 24,
            "Category / Type": 20, "Name": 24, "Transfer": 10,
            "Bank": 18, "Beneficiary": 18, "IBAN / Account": 26, "Amount": 15,
        }
        for i, h in enumerate(base, 1):
            ws.column_dimensions[gc(i)].width = widths.get(h, 14)

        # Agrupar por tanda conservando el orden de la consulta.
        batches: dict[int, dict[str, Any]] = {}
        order: list[int] = []
        for row in rows:
            bid = row["id"]
            if bid not in batches:
                batches[bid] = {"head": row, "lines": []}
                order.append(bid)
            if row["line_no"] is not None:
                batches[bid]["lines"].append(row)

        LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
        RIGHT = Alignment(horizontal="right", vertical="center")

        r = 1
        for bid in order:
            head = batches[bid]["head"]
            lines = batches[bid]["lines"]

            if head["send_date"]:
                d = head["send_date"]
                date_str = f"{MESES[d.month]} {d.day}, {d.year}"
            else:
                d = head["period_month"]
                date_str = f"{MESES[d.month]} {d.year}"
            status_label = STATUS_LABEL.get(head["status"], head["status"])
            title = (
                f"To be sent {date_str}   ·   "
                f"Disbursement #{head['disb_no']}.{head['disb_sub']} · {status_label}"
            )

            # --- Barra de título (navy en el app -> HDR + fuente blanca) ---
            sr = r
            for c in range(1, monto_col + 1):
                ws.cell(sr, c).fill = HDR
            ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=monto_col - 1)
            tc = ws.cell(sr, 1, title)
            tc.font = WHITE
            tc.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[sr].height = 18

            # --- Encabezados de columna ---
            hr = sr + 1
            for i, h in enumerate(base, 1):
                c = ws.cell(hr, i, h)
                c.fill = SUBHDR
                c.font = BOLD
                c.alignment = RIGHT if i == monto_col else LEFT

            # --- Líneas del desembolso (datos base = valores) ---
            dr = hr + 1
            rr = dr
            for ln in lines:
                ws.cell(rr, 1, ln["line_no"])
                ws.cell(rr, 2, (ln["wbs_code"] or "")).font = Font(name="Consolas", size=10)
                ws.cell(rr, 3, ln["description"])
                ws.cell(rr, 4, ln["invoice_no"])
                ws.cell(rr, 5, ln["reason"])
                # Categoría / Tipo en una sola celda (categoría arriba, tipo abajo)
                cat = ln["category"] or ""
                typ = ln["type"] or ""
                ct_txt = f"{cat}\n{typ}" if cat and typ else cat or typ
                ct = ws.cell(rr, 6, ct_txt)
                ct.alignment = Alignment(wrap_text=True, vertical="top")
                ws.cell(rr, 7, ln["payee_name"])
                ws.cell(rr, 8, ln["transfer"])
                col = 9
                if show_bank:
                    ws.cell(rr, col, ln["bank_name"])  # Banco
                    ws.cell(rr, col + 1, ln["legal_id"])  # Beneficiario
                    ws.cell(rr, col + 2, ln["iban"])  # IBAN / Cuenta
                    col += 3
                mc = ws.cell(rr, monto_col, float(ln["amount"] or 0))
                mc.number_format = MONEY
                mc.alignment = RIGHT
                rr += 1
            last = rr - 1

            # --- Fila TOTAL (SUM viva) ---
            tr = rr
            for c in range(1, monto_col + 1):
                ws.cell(tr, c).fill = TOTAL
            ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=monto_col - 1)
            tl = ws.cell(tr, 1, "Total")
            tl.font = BOLD
            tl.alignment = Alignment(horizontal="left", vertical="center")
            if lines:
                total_cell = ws.cell(tr, monto_col, f"=SUM({col_letter}{dr}:{col_letter}{last})")
            else:
                total_cell = ws.cell(tr, monto_col, 0)
            total_cell.number_format = MONEY
            total_cell.font = BOLD
            total_cell.alignment = RIGHT

            # Total de la barra de título = referencia VIVA a la suma del pie.
            ht = ws.cell(sr, monto_col, f"={col_letter}{tr}")
            ht.number_format = MONEY
            ht.font = WHITE
            ht.alignment = RIGHT

            # Siguiente tanda: fila en blanco de separación.
            r = tr + 2

        ws.sheet_view.showGridLines = False

    # ------------------------------------------------------------------ 5
    def _sheet_creditos(self) -> None:
        """Estado de cuenta de crédito — RÉPLICA de CreditView.tsx, FORMULADA.

        Encabezado con Acumulado / Aplicado / Disponible (fórmulas vivas sobre
        la tabla) + estado de cuenta con SALDO CORRIDO por fórmula
        (saldo fila anterior + movimiento de la fila), igual que v_credit_ledger.

        Fuente: v_credit_ledger (move_date, kind, description, amount) — que a su
        vez une accrual (ledger liberado), adjustment (credit_adjustment) y
        application (credit_application). El saldo y los KPIs del header son
        FÓRMULA, no valores pegados.
        """
        ws = self.wb.create_sheet("Credits")
        DATEF = "yyyy-mm-dd"

        # Etiquetas de movimiento — mismas que KIND_LABEL en el front
        KIND_LABEL = {
            "accrual": "Accrual (released ledger)",
            "adjustment": "Manual adjustment",
            "application": "Applied to disbursement",
        }
        APP_LABEL = KIND_LABEL["application"]  # ancla para el SUMIF de "Aplicado"

        ws["A1"] = "DASHBOARD VENTANAS — Credit Account Statement"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = "Live model. Running balance and KPIs by formula."
        ws["A2"].font = Font(italic=True, color="FF808080")

        # --- Tabla: estado de cuenta (saldo corrido) --------------------
        heads = ["Date", "Movement", "Detail", "Amount", "Balance"]
        hr = 7
        for i, h in enumerate(heads, 1):
            c = ws.cell(hr, i, h)
            c.fill = HDR
            c.font = WHITE
            c.alignment = Alignment(horizontal="center")

        rows = self.db.execute(
            text(
                """
                SELECT move_date, kind, description, amount
                FROM v_credit_ledger
                """
            )
        ).mappings().all()

        first = hr + 1
        r = first
        for e in rows:
            ws.cell(r, 1, e["move_date"]).number_format = DATEF
            ws.cell(r, 2, KIND_LABEL.get(e["kind"], e["kind"]))
            ws.cell(r, 3, e["description"])
            ws.cell(r, 4, float(e["amount"] or 0)).number_format = MONEY
            # SALDO CORRIDO vivo: primera fila = su monto; resto = saldo previo + monto
            if r == first:
                ws.cell(r, 5, f"=D{r}").number_format = MONEY
            else:
                ws.cell(r, 5, f"=E{r - 1}+D{r}").number_format = MONEY
            ws.cell(r, 5).font = BOLD
            r += 1
        last = r - 1
        has_rows = last >= first

        if not has_rows:
            ws.cell(first, 1, "No credit movements yet.")
            ws.merge_cells(start_row=first, start_column=1, end_row=first, end_column=5)
            ws.cell(first, 1).alignment = Alignment(horizontal="center")
            ws.cell(first, 1).font = Font(italic=True, color="FF808080")

        # --- Encabezado KPIs: Acumulado / Aplicado / Disponible ---------
        # Fórmulas vivas (replican v_credit_balance):
        #   aplicado   = -SUM(amount) FILTER (kind = 'application')
        #   acumulado  =  SUM(amount) FILTER (kind <> 'application')
        #   disponible =  SUM(amount)  (= último saldo corrido)
        if has_rows:
            amt = f"$D${first}:$D${last}"
            mov = f"$B${first}:$B${last}"
            f_aplicado = f'=-SUMIF({mov},"{APP_LABEL}",{amt})'
            f_acumulado = f'=SUM({amt})-SUMIF({mov},"{APP_LABEL}",{amt})'
            f_disponible = f"=SUM({amt})"
        else:
            f_aplicado = f_acumulado = f_disponible = "0"

        cards = [
            (1, "Accrued", f_acumulado),
            (2, "Applied", f_aplicado),
            (3, "Available", f_disponible),
        ]
        for col, label, formula in cards:
            lc = ws.cell(4, col, label)
            lc.font = Font(bold=True, color="FF808080", size=9)
            lc.fill = SUBHDR
            lc.alignment = Alignment(horizontal="center")
            vc = ws.cell(5, col, formula)
            vc.number_format = MONEY
            vc.font = Font(bold=True, size=13)
            vc.fill = TOTAL
            vc.alignment = Alignment(horizontal="center")

        # anchos y freeze
        for c, w in zip(range(1, 6), [14, 30, 52, 16, 16], strict=False):
            ws.column_dimensions[gc(c)].width = w
        ws.freeze_panes = f"A{first}"
        if has_rows:
            ws.auto_filter.ref = f"A{hr}:E{last}"

    def _sheet_bancos(self) -> None:
        """Estado de cuenta por cada bank_account (réplica de BankView →
        AccountStatement) con SALDO ENCADENADO por fórmula viva.

        Réplica exacta de la vista del app:
          - columnas: Fecha · Nº · Descripción · Débito · Crédito · Cuenta · Tipo · Saldo
          - fila "Saldo inicial" (valor base) + movimientos + fila "Saldo final".
        Regla de oro: el saldo NO se pega — cada fila es
            saldo = saldo_anterior + crédito − débito  (fórmula viva).
        El saldo inicial se deriva igual que el front:
            opening = primer_balance − (primer_credito − primer_debito).
        """
        ws = self.wb.create_sheet("Banks")
        ws["A1"] = "DASHBOARD VENTANAS — Banks (Account statement)"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = "Live model: each balance = previous balance + credit − debit."
        ws["A2"].font = Font(italic=True, color="FF808080")

        DATEF = "yyyy-mm-dd"  # formato de fecha local a esta hoja
        headers = [
            "Date", "No.", "Description", "Debit", "Credit", "Account", "Type", "Balance",
        ]

        accounts = self.db.execute(
            text("SELECT id, name FROM bank_account ORDER BY name")
        ).mappings().all()

        r = 4
        for acc in accounts:
            # --- título de la cuenta ---
            for i in range(1, len(headers) + 1):
                ws.cell(r, i).fill = CATF
            ws.cell(r, 1, acc["name"]).font = Font(bold=True, size=12)
            r += 1

            # --- cabecera ---
            for i, h in enumerate(headers, 1):
                c = ws.cell(r, i, h)
                c.fill = HDR
                c.font = WHITE
                c.alignment = Alignment(horizontal="center")
            r += 1

            # Réplica del SELECT statement() del repositorio (bank_tx), más el
            # label del tipo (bank_movement_class) que el front mapea client-side.
            rows = self.db.execute(
                text(
                    """
                    SELECT t.id, t.tx_date, t.txn_no, t.description,
                           t.debit, t.credit, a.name AS account_name,
                           mc.label AS class_label, t.balance
                    FROM bank_tx t
                    LEFT JOIN bank_account a         ON a.id = t.account_id
                    LEFT JOIN bank_movement_class mc ON mc.code = t.class_code
                    WHERE t.account_id = :aid
                    ORDER BY t.tx_date, t.id
                    """
                ),
                {"aid": acc["id"]},
            ).mappings().all()

            if not rows:
                ws.cell(r, 1, "This account has no movements.").font = Font(
                    italic=True, color="FF808080"
                )
                r += 2
                continue

            # --- Saldo inicial (valor base, igual que el app) ---
            first = rows[0]
            net_first = float(first["credit"] or 0) - float(first["debit"] or 0)
            opening = (
                float(first["balance"]) - net_first
                if first["balance"] is not None
                else 0.0
            )
            opening_row = r
            for i in range(1, len(headers) + 1):
                ws.cell(r, i).fill = SUBHDR
            ws.cell(r, 3, "Opening balance").font = BOLD
            oc = ws.cell(r, 8, round(opening, 2))
            oc.number_format = MONEY
            oc.font = BOLD
            r += 1

            # --- movimientos con saldo corrido VIVO ---
            prev_saldo = f"H{opening_row}"
            for e in rows:
                ws.cell(r, 1, e["tx_date"]).number_format = DATEF
                ws.cell(r, 2, e["txn_no"])
                ws.cell(r, 3, e["description"])
                ws.cell(r, 4, float(e["debit"] or 0)).number_format = MONEY
                ws.cell(r, 5, float(e["credit"] or 0)).number_format = MONEY
                ws.cell(r, 6, e["account_name"])
                ws.cell(r, 7, e["class_label"])
                # saldo = saldo anterior + crédito(E) − débito(D)  → fórmula viva
                ws.cell(r, 8, f"={prev_saldo}+E{r}-D{r}").number_format = MONEY
                prev_saldo = f"H{r}"
                r += 1

            last_data_row = r - 1
            # --- Saldo final = último saldo corrido (fórmula viva) ---
            for i in range(1, len(headers) + 1):
                ws.cell(r, i).fill = TOTAL
            ws.cell(r, 3, "Closing balance").font = BOLD
            fc = ws.cell(r, 8, f"=H{last_data_row}")
            fc.number_format = MONEY
            fc.font = BOLD
            r += 2  # línea en blanco entre cuentas

        for c, w in zip(range(1, 9), [12, 10, 46, 14, 14, 22, 20, 16], strict=False):
            ws.column_dimensions[gc(c)].width = w
        ws.freeze_panes = "A4"

    def _sheet_wires(self) -> None:
        """US Wire Transfers — réplica FORMULADA de la vista del app (UsWiresView.tsx).

        Aportes de los dueños al proyecto. Dos columnas de monto:
          - Enviado por dueños  (amount_sent, dato base; en blanco si NULL -> "—")
          - Recibido en LAFISE  (amount_received, dato base editable)
        y un SALDO LAFISE corrido por FÓRMULA VIVA (saldo anterior + recibido),
        continuo a través de los bloques en el orden en que los muestra el app.

        Estructura espejo: bloque "Ventanas I" (escrow) + bloque "Ventanas II"
        (Disbursement Requests), cada uno con su banda de bloque, sus filas y un
        subtotal (recibido) por fórmula; al pie "TOTAL TRANSFER" (enviado /
        recibido / saldo) y la métrica superior referencia esa fila. El "LAFISE
        Balance" (settings.lafise_balance) es un dato base manual al pie.
        """
        ws = self.wb.create_sheet("US Wire Transfers")

        # --- estilos locales de esta hoja (colores del app) ---
        NAVY = PatternFill("solid", fgColor="FF2D3A5C")        # cabecera de tabla (#2d3a5c)
        BLOCK_FILL = PatternFill("solid", fgColor="FFEAF0FA")   # banda de bloque (#EAF0FA)
        SUB_FILL = PatternFill("solid", fgColor="FFDCE6F4")     # fila subtotal (#DCE6F4)
        NOTE_FONT = Font(italic=True, color="FF808080", size=9)
        LBL_FONT = Font(bold=True, color="FF64748B", size=9)
        GREEN = Font(bold=True, color="FF047857")               # "Recibido" en verde (emerald)
        DATEF = "yyyy-mm-dd"
        RIGHT = Alignment(horizontal="right")

        # bloques en el MISMO orden que la vista (constante BLOCKS de UsWiresView.tsx)
        BLOCKS = [
            ("ventanas_i", "Ventanas I",
             "Contributions to escrow (Hovde Master -> Alta Batalla Escrow)"),
            ("ventanas_ii", "Ventanas II",
             "Disbursement Requests (Hovde Master -> SunWest -> Escrow)"),
        ]

        # --- datos base (mismo SELECT del router us_wires.list_us_wires) ---
        rows = list(
            self.db.execute(
                text(
                    """
                    SELECT id, block, trans_label, wire_date, sender,
                           from_party, to_party, account_number,
                           amount_received, amount_sent, notas
                    FROM us_wire_row
                    ORDER BY sort_order, id
                    """
                )
            ).mappings()
        )
        lafise_balance = self.db.execute(
            text("SELECT lafise_balance FROM settings LIMIT 1")
        ).scalar()

        # --- encabezado del tab ---
        ws["A1"] = "USA to Costa Rica Wire Transfers"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = "Record of all the money the owners have transferred to the project."
        ws["A2"].font = NOTE_FONT

        # resumen (dos metricas del app): valores = fila TOTAL TRANSFER (fórmula viva)
        ws["A4"] = "Sent by owners"
        ws["A4"].font = LBL_FONT
        ws["C4"] = "Received at LAFISE"
        ws["C4"].font = LBL_FONT
        # A5/C5 se llenan al final (referencian la fila TOTAL)

        # --- cabecera de la tabla (mismas columnas / orden que el app) ---
        headers = [
            "Trans #", "Date", "Sender", "From", "To", "Account",
            "Sent owners", "Received LAFISE", "LAFISE Balance", "Notes",
        ]
        hr = 7
        for i, h in enumerate(headers, 1):
            c = ws.cell(hr, i, h)
            c.fill = NAVY
            c.font = WHITE
            c.alignment = Alignment(
                wrap_text=True, vertical="center", horizontal="center"
            )
        ws.row_dimensions[hr].height = 26

        # --- cuerpo: por bloque, en el orden del app ---
        r = hr + 1
        prev_saldo: str | None = None       # ref a la celda Saldo de la fila anterior
        all_data_rows: list[int] = []       # todas las filas de datos (para TOTAL)

        for key, label, note in BLOCKS:
            # banda de bloque (nombre + nota)
            for i in range(1, len(headers) + 1):
                ws.cell(r, i).fill = BLOCK_FILL
            bc = ws.cell(r, 1, label)
            bc.font = BOLD
            ws.cell(r, 2, "- " + note).font = NOTE_FONT
            ws.merge_cells(start_row=r, start_column=2,
                           end_row=r, end_column=len(headers))
            r += 1

            block_rows: list[int] = []
            for e in [x for x in rows if x["block"] == key]:
                ws.cell(r, 1, e["trans_label"])
                if e["wire_date"] is not None:
                    dc = ws.cell(r, 2, e["wire_date"])
                    dc.number_format = DATEF
                ws.cell(r, 3, e["sender"])
                ws.cell(r, 4, e["from_party"])
                ws.cell(r, 5, e["to_party"])
                ws.cell(r, 6, e["account_number"])
                # Enviado (dato base; en blanco si NULL -> el app muestra "—")
                if e["amount_sent"] is not None:
                    sc = ws.cell(r, 7, float(e["amount_sent"]))
                    sc.number_format = MONEY
                # Recibido en LAFISE (dato base editable)
                rc = ws.cell(r, 8, float(e["amount_received"] or 0))
                rc.number_format = MONEY
                rc.font = GREEN
                # Saldo LAFISE corrido = saldo anterior + recibido (FÓRMULA VIVA)
                formula = f"=H{r}" if prev_saldo is None else f"={prev_saldo}+H{r}"
                sal = ws.cell(r, 9, formula)
                sal.number_format = MONEY
                prev_saldo = f"I{r}"
                ws.cell(r, 10, e["notas"])
                block_rows.append(r)
                all_data_rows.append(r)
                r += 1

            # subtotal del bloque (recibido) — FÓRMULA VIVA, en la columna Recibido
            for i in range(1, len(headers) + 1):
                ws.cell(r, i).fill = SUB_FILL
            slab = ws.cell(r, 1, f"Subtotal {label} (received)")
            slab.font = BOLD
            slab.alignment = RIGHT
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
            sub = ws.cell(
                r, 8,
                ("=" + "+".join(f"H{x}" for x in block_rows)) if block_rows else 0,
            )
            sub.number_format = MONEY
            sub.font = BOLD
            r += 1

        # --- fila TOTAL TRANSFER (enviado / recibido / saldo final) ---
        total_row = r
        for i in range(1, len(headers) + 1):
            ws.cell(r, i).fill = TOTAL
        tl = ws.cell(r, 1, "TOTAL TRANSFER")
        tl.font = BOLD
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        g = ws.cell(
            r, 7,
            ("=" + "+".join(f"G{x}" for x in all_data_rows)) if all_data_rows else 0,
        )
        g.number_format = MONEY
        g.font = BOLD
        h = ws.cell(
            r, 8,
            ("=" + "+".join(f"H{x}" for x in all_data_rows)) if all_data_rows else 0,
        )
        h.number_format = MONEY
        h.font = BOLD
        # Saldo total = ultimo saldo corrido (= total recibido). Si no hay filas, 0.
        it = ws.cell(r, 9, f"={prev_saldo}" if prev_saldo else 0)
        it.number_format = MONEY
        it.font = BOLD
        r += 1

        # --- resumen superior: referencia viva a la fila TOTAL ---
        a5 = ws.cell(5, 1, f"=G{total_row}")
        a5.number_format = MONEY
        a5.font = BOLD
        c5 = ws.cell(5, 3, f"=H{total_row}")
        c5.number_format = MONEY
        c5.font = GREEN

        # --- LAFISE Balance (dato base manual, desde settings) ---
        r += 1
        lb = ws.cell(r, 8, "LAFISE Balance")
        lb.font = BOLD
        lb.alignment = RIGHT
        val = ws.cell(
            r, 9, float(lafise_balance) if lafise_balance is not None else None
        )
        val.number_format = MONEY
        val.font = BOLD

        # --- anchos de columna ---
        widths = [20, 12, 16, 26, 26, 18, 16, 16, 16, 30]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[gc(i)].width = w
        ws.freeze_panes = ws.cell(hr + 1, 1)

    # ------------------------------------------------------------------ 5
    def _sheet_facturas(self) -> None:
        """Listado de facturas con sus líneas + TOTAL por fórmula (réplica de InvoicesView).

        Reproduce la vista del app: la LISTA (Nº / Proveedor / Emisión / Total) y el
        DETALLE de cada factura (líneas + Subtotal / Impuesto / Total). Modelo vivo,
        igual que services/invoice.py::_recompute:
          - Subtotal = SUM de los montos de sus líneas  (fórmula)
          - Total    = Subtotal + Impuesto              (fórmula)
        Las anuladas (status='void') quedan en el libro, marcadas ANULADA y tachadas,
        pero NO se borran ni se excluyen del listado (solo no suman al gran total).
        """
        ws = self.wb.create_sheet("Invoices")
        DATEF = "yyyy-mm-dd"
        STRIKE = Font(strike=True)
        VOIDF = Font(strike=True, color="FF999999")
        VOID_FILL = PatternFill("solid", fgColor="FFF4CCCC")

        ws["A1"] = "DASHBOARD VENTANAS — Invoices"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = "Live model. Subtotal = Σ lines · Total = Subtotal + Tax."
        ws["A2"].font = Font(italic=True, color="FF808080")

        cols = [
            "Invoice #", "Supplier", "Issue", "Due", "Status",
            "Description", "WBS (allocation)", "Amount",
        ]
        hr = 3
        for i, h in enumerate(cols, 1):
            c = ws.cell(hr, i, h)
            c.fill = HDR
            c.font = WHITE
        ws.row_dimensions[hr].height = 18

        # Cabeceras — orden IDÉNTICO a list_invoices (Invoice.issue_date.desc())
        invoices = self.db.execute(
            text(
                """
                SELECT i.id, i.invoice_no, i.issue_date, i.due_date,
                       i.tax, i.status, p.name AS payee
                FROM invoice i
                LEFT JOIN payee p ON p.id = i.payee_id
                ORDER BY i.issue_date DESC, i.id DESC
                """
            )
        ).mappings().all()

        # Líneas por factura (con el código WBS del reparto), en orden line_no
        lines_by_inv: dict[int, list[RowMapping]] = defaultdict(list)
        for ln in self.db.execute(
            text(
                """
                SELECT l.invoice_id, l.line_no, l.description, l.amount, w.wbs_code
                FROM invoice_line l
                LEFT JOIN wbs_item w ON w.id = l.wbs_id
                ORDER BY l.invoice_id, l.line_no
                """
            )
        ).mappings():
            lines_by_inv[ln["invoice_id"]].append(ln)

        r = hr + 1
        active_totals: list[int] = []  # filas Total de facturas NO anuladas
        for inv in invoices:
            voided = inv["status"] == "void"

            # --- fila cabecera de la factura (campos de la LISTA del app) ---
            ws.cell(r, 1, inv["invoice_no"]).font = VOIDF if voided else BOLD
            ws.cell(r, 2, inv["payee"]).font = STRIKE if voided else Font()
            ws.cell(r, 3, inv["issue_date"]).number_format = DATEF
            if inv["due_date"]:
                ws.cell(r, 4, inv["due_date"]).number_format = DATEF
            est = ws.cell(r, 5, "VOID" if voided else "Active")
            if voided:
                est.fill = VOID_FILL
                est.font = Font(bold=True, color="FF9C0006")
            else:
                for i in range(1, len(cols) + 1):
                    ws.cell(r, i).fill = SUBHDR
            r += 1

            # --- líneas (valores BASE) ---
            lines = lines_by_inv.get(inv["id"], [])
            first_line = r
            for ln in lines:
                ws.cell(r, 6, ln["description"])
                ws.cell(r, 7, (ln["wbs_code"] or "").strip() or "—")
                ws.cell(r, 8, float(ln["amount"] or 0)).number_format = MONEY
                r += 1
            last_line = r - 1

            # --- Subtotal (Σ líneas) · Impuesto (valor) · Total (fórmula) ---
            ws.cell(r, 6, "Subtotal").font = BOLD
            if lines:
                ws.cell(r, 8, f"=SUM(H{first_line}:H{last_line})").number_format = MONEY
            else:
                ws.cell(r, 8, 0).number_format = MONEY
            sub_row = r
            r += 1

            ws.cell(r, 6, "Tax").font = BOLD
            ws.cell(r, 8, float(inv["tax"] or 0)).number_format = MONEY
            tax_row = r
            r += 1

            ws.cell(r, 6, "Total").font = VOIDF if voided else BOLD
            tot = ws.cell(r, 8, f"=H{sub_row}+H{tax_row}")
            tot.number_format = MONEY
            tot.font = VOIDF if voided else BOLD
            for i in range(1, len(cols) + 1):
                ws.cell(r, i).fill = TOTAL
            if not voided:
                active_totals.append(r)
            r += 2  # una fila en blanco entre facturas

        # --- Gran total (fórmula) — solo facturas ACTIVAS; las anuladas no suman ---
        for i in range(1, len(cols) + 1):
            ws.cell(r, i).fill = HDR
            ws.cell(r, i).font = WHITE
        ws.cell(r, 6, "TOTAL ACTIVE INVOICES")
        if active_totals:
            ws.cell(
                r, 8, "=" + "+".join(f"H{t}" for t in active_totals)
            ).number_format = MONEY

        for c, w in zip(range(1, 9), [16, 26, 12, 12, 12, 46, 18, 14], strict=False):
            ws.column_dimensions[gc(c)].width = w
        ws.freeze_panes = "A4"

    def _sheet_legend(self) -> None:
        ws = self.wb.create_sheet("Legend")
        ws["B2"] = "Status legend"
        ws["B2"].font = Font(bold=True, size=12)
        r = 4
        for code, label, hexc in STATES:
            ws.cell(r, 2).fill = PatternFill("solid", fgColor="FF" + hexc)
            ws.cell(r, 3, label).font = BOLD
            ws.cell(r, 4, code)
            r += 1
        ws["B11"] = "How to collapse"
        ws["B11"].font = Font(bold=True, size=12)
        for i, t in enumerate(
            [
                "The 1 · 2 · 3 buttons at the top left of Job Cost Report:",
                "   1 → quarters only",
                "   2 → quarters + months",
                "   3 → full weekly detail",
                "The +/− above each column group also work.",
                "",
                "LIVE model: Spend, Remaining, Forecast, subtotals and totals are",
                "formulas. If you edit the LEDGER, everything recalculates on its own.",
            ]
        ):
            ws.cell(13 + i, 2, t)
        ws.column_dimensions["B"].width = 6
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 18

def build_excel(db: Session, start: dt.date | None = None, end: dt.date | None = None) -> bytes:
    lo, hi = _week_bounds(db, start, end)
    return ExcelExporter(db, lo, hi).build()
