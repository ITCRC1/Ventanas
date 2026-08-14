"""Short Payment ⇄ Excel: bajar las líneas de un desembolso a .xlsx, editarlas en
Excel y volver a subirlas. Columnas = los campos editables por línea.

El archivo es autoexplicativo (encabezados claros); al subir se reemplazan las
líneas del desembolso (debe estar en borrador). El beneficiario (Nombre) se
resuelve por nombre contra el catálogo de payees (si no existe, se crea).
"""

from __future__ import annotations

import io
from decimal import Decimal, InvalidOperation
from typing import Any

# Encabezados de columna (orden fijo). Deben coincidir export ⇄ import.
COLS = ["Description", "Invoice #", "Note", "Name (Beneficiary)", "Transfer", "Amount"]


def lines_to_xlsx(disb: Any, rows: list[dict[str, Any]]) -> bytes:
    """Genera el .xlsx editable con las líneas actuales del desembolso."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Short Payment"

    period = disb.period_month.strftime("%B %Y") if disb.period_month else ""
    ws["A1"] = f"Short Payment — Disbursement #{disb.disb_no}.{disb.disb_sub} · {period}"
    ws["A1"].font = Font(bold=True, size=13, color="1A237E")
    ws["A2"] = "Edit the rows and upload this file again. Do not change the headers."
    ws["A2"].font = Font(italic=True, size=9, color="666666")

    header_row = 4
    navy = PatternFill("solid", fgColor="1A237E")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    for c, name in enumerate(COLS, start=1):
        cell = ws.cell(row=header_row, column=c, value=name)
        cell.fill = navy
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")

    r = header_row + 1
    for row in rows:
        ws.cell(row=r, column=1, value=row.get("description") or "")
        ws.cell(row=r, column=2, value=row.get("invoice_no") or "")
        ws.cell(row=r, column=3, value=row.get("reason") or "")
        ws.cell(row=r, column=4, value=row.get("payee_name") or "")
        ws.cell(row=r, column=5, value=(row.get("transfer") or "SEND"))
        amt = ws.cell(row=r, column=6, value=float(row.get("amount") or 0))
        amt.number_format = "#,##0.00"
        r += 1

    widths = [42, 14, 16, 34, 10, 14]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = f"A{header_row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_lines_xlsx(data: bytes) -> list[dict[str, Any]]:
    """Lee el .xlsx subido y devuelve las líneas {description, invoice_no, reason,
    payee_name, transfer, amount}. Ignora filas vacías. Tolera el encabezado en
    cualquier fila (busca la fila que tenga 'Descripción')."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb.active
    rows_iter = list(ws.iter_rows(values_only=True))
    # Ubicar la fila de encabezado (la que contiene "Descripción").
    header_idx = None
    for i, r in enumerate(rows_iter):
        cells = [str(x).strip().lower() if x is not None else "" for x in r]
        if any(c.startswith("descrip") for c in cells):
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("Header row not found (column 'Description') in the Excel file.")

    header = [str(x).strip().lower() if x is not None else "" for x in rows_iter[header_idx]]

    def col(*keys: str) -> int | None:
        for j, h in enumerate(header):
            if any(h.startswith(k) for k in keys):
                return j
        return None

    ci = {
        "description": col("descrip"),
        "invoice_no": col("factura", "invoice"),
        "reason": col("nota", "note", "razón", "razon", "reason"),
        "payee_name": col("nombre", "name", "benefic"),
        "transfer": col("transfer"),
        "amount": col("monto", "amount"),
    }

    def get(row: tuple[Any, ...], key: str) -> Any:
        j = ci[key]
        return row[j] if j is not None and j < len(row) else None

    out: list[dict[str, Any]] = []
    for r in rows_iter[header_idx + 1 :]:
        desc = get(r, "description")
        amt_raw = get(r, "amount")
        if (desc is None or str(desc).strip() == "") and (amt_raw is None or str(amt_raw).strip() == ""):
            continue  # fila vacía
        try:
            amount = Decimal(str(amt_raw).replace(",", "").replace("$", "").strip() or "0")
        except (InvalidOperation, ValueError):
            amount = Decimal(0)
        transfer = (str(get(r, "transfer")).strip().upper() or "SEND") if get(r, "transfer") else "SEND"
        out.append(
            {
                "description": str(desc).strip() if desc else "",
                "invoice_no": (str(get(r, "invoice_no")).strip() or None)
                if get(r, "invoice_no")
                else None,
                "reason": (str(get(r, "reason")).strip() or None) if get(r, "reason") else None,
                "payee_name": (str(get(r, "payee_name")).strip() or None)
                if get(r, "payee_name")
                else None,
                "transfer": transfer if transfer in ("SEND", "HOLD") else "SEND",
                "amount": amount,
            }
        )
    return out
