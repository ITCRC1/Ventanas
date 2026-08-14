"""VENTANAS: extractor Excel -> modelo relacional normalizado (SQLite + CSV)."""
import sqlite3, csv, os, re, datetime as dt
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as gc

import sys
SRC=sys.argv[1] if len(sys.argv)>1 else '/mnt/user-data/uploads/VENTANAS_JOB_COST_REP_2026___01-08-2026__FINAL.xlsm'
OUT='/home/claude/ventanas/out_ago'
os.makedirs(OUT, exist_ok=True)

wbf=load_workbook(SRC, data_only=False)
wbv=load_workbook(SRC, data_only=True)

def norm(s):
    if s is None: return None
    s=str(s).strip()
    return s if s else None

def num(v):
    if v is None: return None
    if isinstance(v,(int,float)): return float(v)
    try: return float(str(v).replace(',',''))
    except: return None

con=sqlite3.connect(f'{OUT}/ventanas.db')
con.executescript(open('/dev/stdin').read() if False else """
DROP TABLE IF EXISTS wbs; DROP TABLE IF EXISTS ledger; DROP TABLE IF EXISTS schedule_week;
DROP TABLE IF EXISTS bank_tx; DROP TABLE IF EXISTS change_log; DROP TABLE IF EXISTS wire;
CREATE TABLE wbs(
  wbs TEXT PRIMARY KEY, task_title TEXT, task_owner TEXT, category TEXT, phase TEXT,
  status TEXT, budget_original REAL, budget_change REAL, budget_revised REAL,
  spend REAL, remaining REAL, forecast REAL, variance REAL, src_row INT);
CREATE TABLE ledger(
  id INTEGER PRIMARY KEY, wbs TEXT, account TEXT, date TEXT, invoice TEXT, payee TEXT,
  description TEXT, colones REAL, amount REAL, amount_paid REAL, amount_due REAL,
  funding_source TEXT, src_row INT);
CREATE TABLE schedule_week(
  id INTEGER PRIMARY KEY, wbs TEXT, week_start TEXT, planned_amount REAL,
  is_painted INT, fill_rgb TEXT, src_col TEXT);
CREATE TABLE bank_tx(
  id INTEGER PRIMARY KEY, account TEXT, date TEXT, txn_no TEXT, description TEXT,
  debit REAL, credit REAL, balance REAL, src_sheet TEXT, src_row INT);
CREATE TABLE change_log(
  id INTEGER PRIMARY KEY, ts TEXT, sheet TEXT, row INT, old_value TEXT, new_value TEXT);
CREATE TABLE wire(
  id INTEGER PRIMARY KEY, date TEXT, sender TEXT, dest TEXT, amount REAL, note TEXT);
""")

# ---------- 1. WBS (Job Cost Report B12:M...) ----------
jc=wbv['Job Cost Report']
COL={'status':1,'wbs':2,'title':3,'owner':4,'cat':5,'phase':6,
     'bo':7,'bc':8,'br':9,'spend':10,'rem':11,'fc':12,'var':13}
wbs_rows=[]
for r in range(12, jc.max_row+1):
    w=norm(jc.cell(r,COL['wbs']).value); t=norm(jc.cell(r,COL['title']).value)
    if w is None: continue
    if t and t.upper() in ('TOTAL','TOTALS'): continue
    wbs_rows.append((str(w), t, norm(jc.cell(r,COL['owner']).value),
        norm(jc.cell(r,COL['cat']).value), norm(jc.cell(r,COL['phase']).value),
        norm(jc.cell(r,COL['status']).value),
        num(jc.cell(r,COL['bo']).value), num(jc.cell(r,COL['bc']).value),
        num(jc.cell(r,COL['br']).value), num(jc.cell(r,COL['spend']).value),
        num(jc.cell(r,COL['rem']).value), num(jc.cell(r,COL['fc']).value),
        num(jc.cell(r,COL['var']).value), r))
con.executemany("INSERT OR REPLACE INTO wbs VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)", wbs_rows)

# ---------- 2. LEDGER (A6 headers, data from row 7) ----------
lg=wbv['LEDGER']
led=[]
for r in range(7, lg.max_row+1):
    cc=norm(lg.cell(r,1).value)
    if cc is None: continue
    d=lg.cell(r,3).value
    led.append((None, str(cc), norm(lg.cell(r,2).value),
        d.isoformat() if isinstance(d,dt.datetime) else norm(d),
        norm(lg.cell(r,4).value), norm(lg.cell(r,5).value), norm(lg.cell(r,6).value),
        num(lg.cell(r,7).value), num(lg.cell(r,8).value), num(lg.cell(r,9).value),
        num(lg.cell(r,10).value), norm(lg.cell(r,13).value), r))
con.executemany("INSERT INTO ledger VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)", led)

# ---------- 3. SCHEDULE: semanas del Job Cost Report (CD..GJ) + color ----------
jcf=wbf['Job Cost Report']
week_cols={}
for c in range(21, 200):                       # U.. en adelante = semanas
    h=jc.cell(5,c).value
    if isinstance(h, dt.datetime): week_cols[c]=h.date()
    elif isinstance(h,str):
        m=re.match(r'^(\d{2})/(\d{2})/(\d{4})$', h.strip())
        if m: week_cols[c]=dt.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
sched=[]
wbs_by_row={r[13]: r[0] for r in wbs_rows}
for row, w in wbs_by_row.items():
    for c, wk in week_cols.items():
        val=num(jc.cell(row,c).value)
        cell=jcf.cell(row,c)
        rgb=None; painted=0
        f=cell.fill
        if f and f.fill_type=='solid' and f.fgColor and f.fgColor.type=='rgb':
            rgb=f.fgColor.rgb
            if rgb not in ('00000000','FFFFFFFF'): painted=1
        if val or painted:
            sched.append((None, w, wk.isoformat(), val, painted, rgb, gc(c)))
con.executemany("INSERT INTO schedule_week VALUES(?,?,?,?,?,?,?)", sched)

# ---------- 4. BANCOS ----------
bank=[]
for sh, acc in [('Ventanas II','PV2'), ('VENTANAS I','PV1'), ('Ventanas II (2)','PV2-alt')]:
    s=wbv[sh]
    for r in range(6, s.max_row+1):
        d=s.cell(r,2).value
        if not isinstance(d, dt.datetime): continue
        bank.append((None, acc, d.date().isoformat(), norm(s.cell(r,3).value),
            norm(s.cell(r,4).value), num(s.cell(r,5).value), num(s.cell(r,6).value),
            num(s.cell(r,7).value), sh, r))
for sh in ['Bank Statement LAFISE','Bank Statement LAFISE (2)']:
    s=wbv[sh]
    for r in range(2, s.max_row+1):
        d=s.cell(r,1).value
        if not isinstance(d, dt.datetime): continue
        bank.append((None,'LAFISE', d.date().isoformat(), None,
            norm(s.cell(r,2).value), num(s.cell(r,3).value), num(s.cell(r,4).value),
            num(s.cell(r,8).value), sh, r))
con.executemany("INSERT INTO bank_tx VALUES(?,?,?,?,?,?,?,?,?,?)", bank)

# ---------- 5. LOG de Cambios ----------
lc=wbv['LOG de Cambios']; logs=[]
for r in range(2, lc.max_row+1):
    ts=lc.cell(r,1).value
    if ts is None: continue
    logs.append((None, str(ts), norm(lc.cell(r,2).value), int(lc.cell(r,3).value or 0),
                 norm(lc.cell(r,4).value), norm(lc.cell(r,5).value)))
con.executemany("INSERT INTO change_log VALUES(?,?,?,?,?,?)", logs)

# ---------- 6. Wires ----------
uw=wbv['US Wire Transfers']; wires=[]
for r in range(6, uw.max_row+1):
    d=uw.cell(r,1).value
    if not isinstance(d, dt.datetime): continue
    wires.append((None, d.date().isoformat(), norm(uw.cell(r,3).value),
                  norm(uw.cell(r,5).value), num(uw.cell(r,7).value), norm(uw.cell(r,8).value)))
con.executemany("INSERT INTO wire VALUES(?,?,?,?,?,?)", wires)
con.commit()

# ---------- export CSV ----------
for t in ['wbs','ledger','schedule_week','bank_tx','change_log','wire']:
    cur=con.execute(f"SELECT * FROM {t}")
    with open(f'{OUT}/{t}.csv','w',newline='',encoding='utf-8') as fh:
        w=csv.writer(fh); w.writerow([d[0] for d in cur.description]); w.writerows(cur)
    print(f"{t:15s} {con.execute(f'SELECT COUNT(*) FROM {t}').fetchone()[0]:6d} filas")
con.close()
