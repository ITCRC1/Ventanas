#!/usr/bin/env python3
"""Importa la hoja 'Short Term Payments' del Excel a desembolsos.

Cada tanda ("To be sent <fecha>") -> un disbursement; cada pago -> disbursement_line
con el detalle bancario (payee = Name + Bank + Beneficiary + IBAN). El total lo
calcula el trigger de la base.

Uso:
  python import_short_payments.py --xlsx "<ruta>.xlsm" --pg "postgresql://..."
"""
from __future__ import annotations

import argparse
import re
import sys
from datetime import date
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

try:
    import psycopg
except ImportError:
    sys.exit("Falta psycopg")

MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "abr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}
PROPS = {"cinco ventanas lot", "ventanas", "cinco ventanas lot 107", "procore"}
# Acepta ordinal (1st/2nd/3rd/th) y meses en inglés o español (abril).
_DATE = re.compile(
    r"(jan|feb|mar|apr|abr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\s+"
    r"(\d{1,2})(?:st|nd|rd|th)?\.?,?\s+(\d{4})",
    re.I,
)


def parse_date(text: str) -> date | None:
    m = _DATE.search(text or "")
    if not m:
        return None
    mo = MONTHS[m.group(1).lower()[:3]]
    try:
        return date(int(m.group(3)), mo, int(m.group(2)))
    except ValueError:
        return None


def parse_amount(v) -> Decimal | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return Decimal(str(round(float(v), 2)))
    s = str(v).replace("$", "").replace(",", "").strip()
    try:
        return Decimal(s).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--pg", required=True)
    ap.add_argument("--mask-iban", action="store_true", help="no cargar los IBAN (privacidad)")
    a = ap.parse_args()

    wb = load_workbook(a.xlsx, data_only=True)
    ws = wb["Short Term Payments"]
    pg = psycopg.connect(a.pg)
    cur = pg.cursor()
    cur.execute("SELECT set_config('app.user_id', 'import-short-payments', false)")
    cur.execute("SELECT COALESCE(MAX(disb_no),0) FROM disbursement WHERE disb_sub=0")
    disb_no = int(cur.fetchone()[0])

    payee_cache: dict[str, int] = {}

    def get_payee(name, bank, legal, iban) -> int | None:
        if not name:
            return None
        key = name.strip()
        if key in payee_cache:
            return payee_cache[key]
        cur.execute("SELECT id FROM payee WHERE name=%s", (key,))
        row = cur.fetchone()
        if row:
            payee_cache[key] = row[0]
            return row[0]
        cur.execute(
            "INSERT INTO payee (name,bank_name,legal_id,iban) VALUES (%s,%s,%s,%s) RETURNING id",
            (key, bank, legal, iban),
        )
        pid = cur.fetchone()[0]
        payee_cache[key] = pid
        return pid

    batches = lines = 0
    cur_disb = None
    line_no = 0

    def close_batch():
        nonlocal cur_disb, line_no
        cur_disb = None
        line_no = 0

    # La hoja está corrida una columna: descripción en col 2, monto en col 3, etc.
    for r in range(1, ws.max_row + 1):
        c1 = ws.cell(r, 2).value
        t1 = str(c1).strip() if c1 is not None else ""
        low = t1.lower()
        if not t1:
            continue
        # ¿inicio de tanda? (texto con fecha, típicamente "To be sent ..." o "before ...")
        d = parse_date(t1)
        if d and ("sent" in low or "before" in low or "payment" in low or len(t1) < 40):
            close_batch()
            # Solo se agregan los meses que NO existen ya (aditivo, idempotente).
            cur.execute("SELECT 1 FROM disbursement WHERE period_month=%s", (d.replace(day=1),))
            if cur.fetchone():
                continue  # tanda ya cargada -> cur_disb queda None, sus líneas se ignoran
            disb_no += 1
            cur.execute(
                """INSERT INTO disbursement (disb_no,disb_sub,period_month,send_date,status,notes)
                   VALUES (%s,0,%s,%s,'draft',%s) RETURNING id""",
                (disb_no, d.replace(day=1), d, "Importado de Short Term Payments"),
            )
            cur_disb = cur.fetchone()[0]
            line_no = 0
            batches += 1
            continue
        if low in PROPS:  # sub-encabezado de propiedad
            continue
        if low.startswith("total"):
            close_batch()
            continue
        # línea de pago (columnas corridas: monto col3, nota col4, name col5, ...)
        amt = parse_amount(ws.cell(r, 3).value)
        if cur_disb is None or amt is None:
            continue
        note = ws.cell(r, 4).value
        name = ws.cell(r, 5).value
        transfer = ws.cell(r, 6).value
        bank = ws.cell(r, 7).value
        legal = ws.cell(r, 8).value
        iban = None if a.mask_iban else ws.cell(r, 9).value
        pid = get_payee(name, bank, legal, iban)
        line_no += 1
        cur.execute(
            """INSERT INTO disbursement_line
               (disbursement_id,line_no,description,amount,reason,transfer,payee_id)
               VALUES (%s,%s,%s,%s,%s,%s,%s)""",
            (cur_disb, line_no, t1[:200], amt, str(note)[:200] if note else None,
             str(transfer)[:20] if transfer else None, pid),
        )
        lines += 1

    pg.commit()
    print(f"Tandas: {batches} | Líneas: {lines} | Payees: {len(payee_cache)}")
    pg.close()


if __name__ == "__main__":
    main()
