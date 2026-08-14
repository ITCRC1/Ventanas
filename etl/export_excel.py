#!/usr/bin/env python3
"""
DASHBOARD VENTANAS — Exportador de modelo Excel VIVO

Genera un .xlsx que:
  · se ve igual que el master actual (mismos colores, misma disposición)
  · lleva FÓRMULAS reales, no valores pegados — si editás el LEDGER, todo
    recalcula igual que hoy
  · colapsa de verdad, con botones +/− en tres niveles:
        semana  →  mes  →  trimestre
    (el Excel actual NO agrupa: sólo esconde columnas a mano)

Fuente: la base de la app. Aquí lee el SQLite del extractor para poder
probarse sin Postgres; en producción cambia la consulta, no la lógica.
"""
import sqlite3, datetime as dt, argparse, os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gc
from openpyxl.worksheet.properties import Outline
from openpyxl.formatting.rule import CellIsRule

# Leyenda del master (Job Cost Report, filas 129-133)
STATES = [
    ('not_started', 'Not Started', 'CCCCCC'),
    ('in_process',  'In process',  'FFFF00'),
    ('approved',    'Approved',    '1155CC'),
    ('attention',   'Attention',   'B85B22'),
    ('completed',   'Completed',   '38761D'),
]
STATE_FILL = {c: PatternFill('solid', fgColor='FF' + h) for c, _, h in STATES}
COLOR_OF   = {c: h for c, _, h in STATES}

HDR   = PatternFill('solid', fgColor='FF434343')
SUBHDR= PatternFill('solid', fgColor='FFD9D9D9')
TOTAL = PatternFill('solid', fgColor='FFC8DCF0')
CATF  = PatternFill('solid', fgColor='FFF8F8F8')
WHITE = Font(color='FFFFFFFF', bold=True)
BOLD  = Font(bold=True)
MONEY = '#,##0.00;[Red](#,##0.00)'
THIN  = Border(*[Side('thin', color='FFD0D0D0')] * 4)


def mondays(start, end):
    d = start - dt.timedelta(days=start.weekday())
    while d <= end:
        yield d
        d += dt.timedelta(days=7)


def quarter(d):
    return f"Q{(d.month - 1)//3 + 1} {d.year}"


class Exporter:
    def __init__(self, db, start, end):
        self.cx = sqlite3.connect(db)
        self.cx.row_factory = sqlite3.Row
        self.weeks = list(mondays(start, end))
        self.wb = Workbook()

    def q(self, sql, *a):
        return self.cx.execute(sql, a).fetchall()

    # ------------------------------------------------------------------
    def build(self, path):
        self.sheet_ledger()
        self.sheet_jobcost()
        self.sheet_timeline_detail()
        self.sheet_legend()
        self.wb.move_sheet('Job Cost Report', offset=-2)
        self.wb.save(path)
        return path

    # ------------------------------------------------------------------ 1
    def sheet_ledger(self):
        ws = self.wb.active
        ws.title = 'LEDGER'
        cols = ['Cost Code', 'Date', 'Invoice #', 'Payee', 'Description',
                'Amount', 'Amount Paid', 'Amount Due', 'Funding Source']
        ws.append([]); ws.append([])
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            ws.cell(3, c).fill = HDR
            ws.cell(3, c).font = WHITE
        r = 4
        for e in self.q("SELECT * FROM ledger ORDER BY src_row"):
            ws.cell(r, 1, (e['wbs'] or '').strip())
            ws.cell(r, 2, e['date'])
            ws.cell(r, 3, e['invoice'])
            ws.cell(r, 4, e['payee'])
            ws.cell(r, 5, e['description'])
            ws.cell(r, 6, e['amount'] or 0).number_format = MONEY
            ws.cell(r, 7, e['amount_paid'] or 0).number_format = MONEY
            # Amount Due VIVO
            ws.cell(r, 8, f"=F{r}-G{r}").number_format = MONEY
            ws.cell(r, 9, e['funding_source'])
            r += 1
        self.ledger_last = r - 1
        for c, w in zip(range(1, 10), [12, 12, 14, 26, 46, 14, 14, 14, 18]):
            ws.column_dimensions[gc(c)].width = w
        ws.freeze_panes = 'B4'
        ws.auto_filter.ref = f"A3:I{self.ledger_last}"

    # ------------------------------------------------------------------ 2
    def sheet_jobcost(self):
        """El master. Presupuesto con fórmulas + semanas agrupadas en 3 niveles."""
        ws = self.wb.create_sheet('Job Cost Report')
        ws.sheet_properties.outlinePr = Outline(summaryRight=True, summaryBelow=False,
                                                showOutlineSymbols=True)
        ws['A1'] = 'DASHBOARD VENTANAS — Job Cost Report'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = 'Modelo vivo. Editá el LEDGER y todo recalcula.'
        ws['A2'].font = Font(italic=True, color='FF808080')

        fixed = ['Status', 'WBS', 'Task Title', 'Owner', 'Category', 'Phase',
                 'Budget Original', 'Budget Change', 'Budget Revised',
                 'Spend', 'Remaining', 'Forecast', 'Variance']
        HR = 5                                   # fila de encabezados
        for i, h in enumerate(fixed, 1):
            c = ws.cell(HR, i, h); c.fill = HDR; c.font = WHITE
            c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        ws.row_dimensions[HR].height = 30

        # ---- disposición temporal: semanas + subtotal de mes + subtotal de trimestre
        col = len(fixed) + 1
        self.week_col, self.month_col, self.qtr_col = {}, {}, {}
        by_q = defaultdict(lambda: defaultdict(list))
        for w in self.weeks:
            by_q[quarter(w)][(w.year, w.month)].append(w)

        for qname, months in by_q.items():
            q_first = col
            m_cols = []
            for (yy, mm), wks in months.items():
                w_first = col
                for w in wks:
                    ws.cell(3, col, qname)
                    ws.cell(4, col, dt.date(yy, mm, 1).strftime('%b %Y').upper())
                    cc = ws.cell(HR, col, w.day)
                    cc.fill = SUBHDR; cc.font = Font(bold=True, size=9)
                    cc.alignment = Alignment(horizontal='center')
                    ws.column_dimensions[gc(col)].width = 5.5
                    self.week_col[w] = col
                    col += 1
                # subtotal del mes
                mc = ws.cell(HR, col, dt.date(yy, mm, 1).strftime('%b').upper())
                mc.fill = PatternFill('solid', fgColor='FFA2C4C9'); mc.font = BOLD
                mc.alignment = Alignment(horizontal='center')
                ws.cell(3, col, qname); ws.cell(4, col, 'MES')
                ws.column_dimensions[gc(col)].width = 12
                self.month_col[(yy, mm)] = (col, w_first, col - 1)
                m_cols.append(col)
                # las SEMANAS son nivel 2 (las más internas)
                for k in range(w_first, col):
                    ws.column_dimensions[gc(k)].outlineLevel = 2
                col += 1
            # subtotal del trimestre
            qc = ws.cell(HR, col, qname)
            qc.fill = TOTAL; qc.font = BOLD
            qc.alignment = Alignment(horizontal='center')
            ws.cell(3, col, qname); ws.cell(4, col, 'TRIMESTRE')
            ws.column_dimensions[gc(col)].width = 14
            self.qtr_col[qname] = (col, m_cols)
            # los SUBTOTALES DE MES son nivel 1; el del trimestre queda en 0
            for k in m_cols:
                ws.column_dimensions[gc(k)].outlineLevel = 1
            col += 1
        self.total_col = col
        tc = ws.cell(HR, col, 'TOTAL'); tc.fill = HDR; tc.font = WHITE
        ws.column_dimensions[gc(col)].width = 15

        # ---- filas: agrupadas por categoría, con subtotal por categoría
        rows = self.q("""SELECT * FROM wbs WHERE budget_revised IS NOT NULL
                         ORDER BY COALESCE(category,'zzz'), src_row""")
        sched = defaultdict(dict)
        for s in self.q("SELECT * FROM schedule_week"):
            sched[(s['wbs'] or '').strip()][s['week_start']] = s
        state_of = {'Approved': 'approved', 'In process': 'in_process',
                    'Attention': 'attention', 'Completed': 'completed'}

        r = HR + 1
        cat_rows = defaultdict(list)
        self.row_of = {}
        cur = object()
        for e in rows:
            cat = (e['category'] or 'Sin categoría').strip()
            if cat != cur:
                if cur is not object():
                    r = self._cat_total(ws, r, cur, cat_rows[cur]); r += 1
                cur = cat
                cc = ws.cell(r, 1, cat.upper())
                for i in range(1, self.total_col + 1):
                    ws.cell(r, i).fill = CATF
                cc.font = BOLD
                r += 1
            code = (e['wbs'] or '').strip()
            self.row_of[code] = r
            cat_rows[cat].append(r)
            ws.cell(r, 1, e['status'])
            ws.cell(r, 2, code)
            ws.cell(r, 3, e['task_title'])
            ws.cell(r, 4, e['task_owner'])
            ws.cell(r, 5, cat)
            ws.cell(r, 6, e['phase'])
            ws.cell(r, 7, e['budget_original'] or 0).number_format = MONEY
            ws.cell(r, 8, e['budget_change'] or 0).number_format = MONEY
            # ---- FÓRMULAS VIVAS
            ws.cell(r, 9,  f"=G{r}+H{r}").number_format = MONEY
            ws.cell(r, 10, f"=SUMIF(LEDGER!$A:$A,$B{r},LEDGER!$G:$G)").number_format = MONEY
            ws.cell(r, 11, f"=I{r}-J{r}").number_format = MONEY
            ws.cell(r, 12, f"=J{r}+{gc(self.total_col)}{r}").number_format = MONEY
            ws.cell(r, 13, f"=I{r}-L{r}").number_format = MONEY
            # celdas de semana con su color de estado
            st = state_of.get((e['status'] or '').strip(), 'not_started')
            for w, wc in self.week_col.items():
                s = sched[code].get(w.isoformat())
                cell = ws.cell(r, wc)
                if s and s['planned_amount']:
                    cell.value = round(s['planned_amount'], 2)
                    cell.number_format = '#,##0'
                    cell.fill = STATE_FILL[st]
                elif s and s['is_painted']:
                    cell.fill = STATE_FILL['not_started']
            # subtotales de mes / trimestre / total: FÓRMULAS
            for (yy, mm), (mc, a, b) in self.month_col.items():
                ws.cell(r, mc, f"=SUM({gc(a)}{r}:{gc(b)}{r})").number_format = MONEY
            for qn, (qc2, mcs) in self.qtr_col.items():
                ws.cell(r, qc2, "=" + "+".join(f"{gc(m)}{r}" for m in mcs)).number_format = MONEY
            ws.cell(r, self.total_col,
                    "=" + "+".join(f"{gc(q)}{r}" for q, _ in self.qtr_col.values())
                    ).number_format = MONEY
            r += 1
        r = self._cat_total(ws, r, cur, cat_rows[cur]); r += 1

        # gran total
        for i in range(1, self.total_col + 1):
            ws.cell(r, i).fill = TOTAL
        ws.cell(r, 3, 'PROJECT TOTAL').font = BOLD
        allrows = [x for v in cat_rows.values() for x in v]
        for i in list(range(7, 14)) + [self.total_col] + \
                 [c for c, _, _ in self.month_col.values()] + \
                 [c for c, _ in self.qtr_col.values()]:
            ws.cell(r, i, "=" + "+".join(f"{gc(i)}{x}" for x in allrows)).number_format = MONEY
            ws.cell(r, i).font = BOLD
        self.grand_row = r

        for c, w in zip(range(1, 14), [11, 10, 42, 14, 20, 20, 15, 14, 15, 15, 15, 15, 14]):
            ws.column_dimensions[gc(c)].width = w
        ws.freeze_panes = ws.cell(HR + 1, 7)
        # semáforo de sobregiro sobre Remaining
        ws.conditional_formatting.add(f"K{HR+1}:K{r}",
            CellIsRule(operator='lessThan', formula=['0'],
                       fill=PatternFill('solid', fgColor='FFF4CCCC'), font=Font(color='FF9C0006')))

    def _cat_total(self, ws, r, cat, rows):
        for i in range(1, self.total_col + 1):
            ws.cell(r, i).fill = SUBHDR
        ws.cell(r, 3, f'TOTAL {cat}').font = BOLD
        if rows:
            cols = list(range(7, 14)) + [self.total_col] + \
                   [c for c, _, _ in self.month_col.values()] + \
                   [c for c, _ in self.qtr_col.values()]
            for i in cols:
                ws.cell(r, i, "=" + "+".join(f"{gc(i)}{x}" for x in rows)).number_format = MONEY
                ws.cell(r, i).font = BOLD
        return r

    # ------------------------------------------------------------------ 3
    def sheet_timeline_detail(self):
        """Mismo dato, agregado por WBS. Todo por fórmula contra el master."""
        ws = self.wb.create_sheet('Timeline Detail')
        ws.sheet_properties.outlinePr = Outline(summaryRight=True, summaryBelow=False,
                                                showOutlineSymbols=True)
        heads = ['Project Category', 'Phase', 'WBS NUMBER', 'Task Title',
                 'Budget Revised', 'YTD Expense', 'Remaining', 'Forecast']
        for i, h in enumerate(heads, 1):
            c = ws.cell(3, i, h); c.fill = HDR; c.font = WHITE
            c.alignment = Alignment(wrap_text=True, horizontal='center')
        ws.row_dimensions[3].height = 28
        r = 4
        for e in self.q("""SELECT * FROM wbs WHERE budget_revised IS NOT NULL
                           ORDER BY COALESCE(category,'zzz'), src_row"""):
            code = (e['wbs'] or '').strip()
            ws.cell(r, 1, f"=IFERROR(INDEX('Job Cost Report'!$E:$E,"
                          f"MATCH($C{r},'Job Cost Report'!$B:$B,0)),\"\")")
            ws.cell(r, 2, f"=IFERROR(INDEX('Job Cost Report'!$F:$F,"
                          f"MATCH($C{r},'Job Cost Report'!$B:$B,0)),\"\")")
            ws.cell(r, 3, code)
            ws.cell(r, 4, f"=IFERROR(INDEX('Job Cost Report'!$C:$C,"
                          f"MATCH($C{r},'Job Cost Report'!$B:$B,0)),\"\")")
            for i, srccol in zip(range(5, 9), 'IJKL'):
                ws.cell(r, i, f"=SUMIF('Job Cost Report'!$B:$B,$C{r},"
                              f"'Job Cost Report'!${srccol}:${srccol})").number_format = MONEY
            r += 1
        for c, w in zip(range(1, 9), [24, 20, 12, 42, 15, 15, 15, 15]):
            ws.column_dimensions[gc(c)].width = w
        ws.freeze_panes = 'E4'
        ws.auto_filter.ref = f"A3:H{r-1}"
        # La categoría y la fase se HEREDAN del master por fórmula:
        # nunca más pueden divergir como pasaba en el Excel original.

    # ------------------------------------------------------------------ 4
    def sheet_legend(self):
        ws = self.wb.create_sheet('Leyenda')
        ws['B2'] = 'Leyenda de estados'; ws['B2'].font = Font(bold=True, size=12)
        r = 4
        for code, label, hexc in STATES:
            ws.cell(r, 2).fill = PatternFill('solid', fgColor='FF' + hexc)
            ws.cell(r, 3, label).font = BOLD
            ws.cell(r, 4, code)
            r += 1
        ws['B11'] = 'Cómo colapsar'; ws['B11'].font = Font(bold=True, size=12)
        for i, t in enumerate([
            'Los botones 1 · 2 · 3 arriba a la izquierda de la hoja Job Cost Report:',
            '   1 → sólo trimestres',
            '   2 → trimestres + meses',
            '   3 → todo el detalle semanal',
            'También sirven los +/− sobre cada grupo de columnas.',
            '',
            'Este archivo es un modelo VIVO: las columnas Spend, Remaining,',
            'Forecast, los subtotales y los totales son fórmulas.',
            'Si editás el LEDGER, todo recalcula solo.'], 0):
            ws.cell(13 + i, 2, t)
        ws.column_dimensions['B'].width = 6
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 18


if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('--db', default='out/ventanas.db')
    ap.add_argument('--start', default='2026-01-05')
    ap.add_argument('--end',   default='2026-12-28')
    ap.add_argument('--out',   default='DASHBOARD_VENTANAS_modelo_vivo.xlsx')
    a = ap.parse_args()
    p = Exporter(a.db, dt.date.fromisoformat(a.start), dt.date.fromisoformat(a.end)).build(a.out)
    print("generado:", p, os.path.getsize(p), "bytes")
